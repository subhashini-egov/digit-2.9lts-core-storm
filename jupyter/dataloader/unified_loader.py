"""
Unified PGR Data Loader - Core Module
All code for reading Excel and generating API payloads
Users should not modify this file directly
"""

import pandas as pd
import json
import math
import warnings
import requests
import time
import os
from typing import Dict, List, Any
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

warnings.filterwarnings('ignore')

# Load environment variables from .env file
load_dotenv()


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def clean_nans(obj):
    """
    Recursively clean NaN values and convert dates for JSON serialization
    """
    if isinstance(obj, dict):
        cleaned = {}
        for k, v in obj.items():
            cleaned[k] = clean_nans(v)
        return cleaned
    elif isinstance(obj, list):
        return [clean_nans(item) for item in obj]
    elif isinstance(obj, float) and math.isnan(obj):
        return None  # Convert NaN to null
    elif isinstance(obj, (pd.Timestamp, datetime)):
        return obj.isoformat()  # Convert Timestamp/datetime to ISO string
    elif pd.isna(obj):
        return None  # Handle any other pandas NA types
    else:
        return obj  # Keep None as None (will be null in JSON)


# ============================================================================
# EXCEL READER CLASS
# ============================================================================

class UnifiedExcelReader:
    """Read unified Excel template and generate API payloads"""

    def __init__(self, excel_file: str):
        self.excel_file = excel_file

    # ========================================================================
    # TENANT MASTER READERS (PHASE 1)
    # ========================================================================

    def read_tenant_info(self):
        """
        Reads the NEW Tenant Master template.
        Updated to match:
        - Tenant Display Name*
        - Tenant Code*
        - Tenant Type*
        - Logo File Path*
        - Latitude, Longitude
        - City Name
        - District Name
        - Address
        - Tenant Website
        """

        df = pd.read_excel(self.excel_file, sheet_name='Tenant Info')

        tenants = []
        localizations = []
        district_counter = {}  # For auto-generating district codes
        city_counter = {}   

        for _, row in df.iterrows():
            
            # Skip empty rows
            if pd.isna(row.get('Tenant Display Name*')) or pd.isna(row.get('Tenant Code*\n(To be filled by ADMIN)')):
                continue

            tenant_name = str(row['Tenant Display Name*']).strip()
            tenant_code_col = 'Tenant Code*\n(To be filled by ADMIN)'
            tenant_code = str(row[tenant_code_col]).strip().lower()

            tenant_type = str(row.get('Tenant Type*', '')).strip()
            logo_path = str(row.get('Logo File Path*', '')).strip()

            city_name = str(row.get('City Name', '')).strip()
            district_name = str(row.get('District Name', '')).strip()
            address = str(row.get('Address', '')).strip()
            website = str(row.get('Tenant Website', '')).strip()

            latitude = float(row['Latitude']) if pd.notna(row.get('Latitude')) else 0.0
            longitude = float(row['Longitude']) if pd.notna(row.get('Longitude')) else 0.0

            # Build city object
            district_code = ""
            if district_name:
                if district_name not in district_counter:
                    district_counter[district_name] = len(district_counter) + 1
                district_code = f"District_{district_counter[district_name]:03d}"
            city_code = ""
            if city_name:
                if city_name not in city_counter:
                    city_counter[city_name] = len(city_counter) + 1
                city_code = f"City_{city_counter[city_name]:03d}"
    
            city = {
               'code': city_code,
                'name': city_name or tenant_name,
                'districtName': district_name,
                'districtTenantCode': district_code,
                'ulbGrade':'',
                'latitude': latitude,
                'longitude': longitude,
                'localName': tenant_name,
                'captcha': 'true'
            }

            # Tenant object
            tenant = {
                'code': tenant_code,
                'name': tenant_name,
                'type': tenant_type,
                'emailId': "",
                'contactNumber': "",
                'address': address,
                'domainUrl': website or "https://example.com",
                'logoId': logo_path,
                'imageId': logo_path,
                'description': tenant_name,
                'twitterUrl': "",
                'facebookUrl': "",
                'OfficeTimings': {'Mon - Fri': '10:00 AM - 5:00 PM'},
                'city': city
            }

            tenants.append(tenant)

            # Add localization
            loc_code = f"TENANT_TENANTS_{tenant_code.upper().replace('.', '_')}"
            localizations.append({
                "code": loc_code,
                "message": tenant_name,
                "module": "rainmaker-common",
                "locale": "en_IN"
            })
        return tenants, localizations


    def read_tenant_branding(self, tenant_id: str):
        """Read Tenant Branding sheet

        Returns:
            list: Branding information records for MDMS upload
        """
        try:
            df = pd.read_excel(self.excel_file, sheet_name='Tenant Branding Details')
        except Exception as e:
            print(f"⚠️ Could not read 'Tenant Branding Details' sheet: {str(e)}")
            return []

        branding_list = []
        for _, row in df.iterrows():
                branding_record = {
                     'code':tenant_id,
                    'name':tenant_id,
                    'bannerUrl': str(row.get('Banner URL', '')) if pd.notna(row.get('Banner URL')) else "",
                    'logoUrl': str(row.get('Logo URL', '')) if pd.notna(row.get('Logo URL')) else "",
                    'logoUrlWhite': str(row.get('Logo URL (White)', '')) if pd.notna(row.get('Logo URL (White)')) else "",
                    'statelogo': str(row.get('State Logo', '')) if pd.notna(row.get('State Logo')) else ""
                }
                branding_list.append(branding_record)

        return branding_list

    # ========================================================================
    # COMMON MASTER READERS (PHASE 2)
    # ========================================================================

    def read_departments_designations(self, tenant_id: str, uploader=None):
        """Read combined Department and Designation sheet from Common Master Excel
        Auto-generates codes for departments and designations
        Fetches existing data from MDMS to continue numbering

        Args:
            tenant_id: Tenant ID for localization context
            uploader: Authenticated APIUploader instance (optional, for fetching existing data)

        Returns:
            tuple: (departments_list, designations_list, dept_localization, desig_localization, dept_name_to_code_mapping)
        """
        df = pd.read_excel(self.excel_file, sheet_name='Department And Desgination Mast')

        departments = []
        designations = []
        dept_localizations = []
        desig_localizations = []

        dept_counter = {}
        dept_name_to_code = {}  # Mapping for complaint types
        desig_counter = 1

        # Fetch existing departments and designations from MDMS to continue numbering
        if uploader:
            try:
                existing_depts = uploader.fetch_departments(tenant_id)
                existing_desigs = uploader.fetch_designations(tenant_id)

                # Find max department counter
                max_dept_num = 0
                for dept in existing_depts:
                    code = dept.get('code', '')
                    if code.startswith('DEPT_'):
                        try:
                            num = int(code.split('_')[1])
                            max_dept_num = max(max_dept_num, num)
                        except (ValueError, IndexError):
                            pass
                    # Map existing dept names to codes
                    dept_name_to_code[dept.get('name', '')] = code

                # Find max designation counter
                max_desig_num = 0
                for desig in existing_desigs:
                    code = desig.get('code', '')
                    if code.startswith('DESIG_'):
                        try:
                            num = int(code.split('_')[1])
                            max_desig_num = max(max_desig_num, num)
                        except (ValueError, IndexError):
                            pass

                # Start counters from next available number
                dept_start_counter = max_dept_num + 1
                desig_counter = max_desig_num + 1

            except Exception as e:
                # If fetch fails, start from 1
                dept_start_counter = 1
                desig_counter = 1
        else:
            dept_start_counter = 1

        for _, row in df.iterrows():
            dept_name = row.get('Department Name*')
            desig_name = row.get('Designation Name*')

            # Skip empty rows
            if pd.isna(dept_name) or str(dept_name).strip() == '':
                continue

            dept_name = str(dept_name).strip()

            # Check if department already exists in MDMS
            if dept_name in dept_name_to_code:
                dept_code = dept_name_to_code[dept_name]
            else:
                # Auto-generate department code with next available number
                if dept_name not in dept_counter:
                    dept_counter[dept_name] = dept_start_counter
                    dept_start_counter += 1
                    dept_code = f"DEPT_{dept_counter[dept_name]}"

                    # Store mapping from name to code
                    dept_name_to_code[dept_name] = dept_code

                    # Add department
                    departments.append({
                        'code': dept_code,
                        'name': dept_name,
                        'active': True
                    })

                    # Auto-generate department localization
                    loc_code = f"COMMON_MASTERS_DEPARTMENT_{dept_code}"
                    dept_localizations.append({
                        'code': loc_code,
                        'message': dept_name,
                        'module': 'rainmaker-common',
                        'locale': 'en_IN'
                    })
                else:
                    dept_code = f"DEPT_{dept_counter[dept_name]}"

            # Add designation if present
            if pd.notna(desig_name) and str(desig_name).strip() != '':
                desig_name = str(desig_name).strip()
                desig_code = f"DESIG_{desig_counter:02d}"
                desig_counter += 1

                designations.append({
                    'code': desig_code,
                    'name': desig_name,
                    'department': [dept_code],
                    'active': True,
                    'description': f'{desig_name} - {dept_name}'
                })

                # Auto-generate designation localization
                loc_code = f"COMMON_MASTERS_{desig_code}"
                desig_localizations.append({
                    'code': loc_code,
                    'message': desig_name,
                    'module': 'rainmaker-common',
                    'locale': 'en_IN'
                })

        return departments, designations, dept_localizations, desig_localizations, dept_name_to_code

    def read_complaint_types(self, tenant_id: str, dept_name_to_code: dict = None):
        """Read Complaint Type Master from Common Master Excel
        Auto-generates service codes and handles hierarchical structure

        Args:
            tenant_id: Tenant ID for context
            dept_name_to_code: Dictionary mapping department names to codes

        Returns:
            tuple: (complaint_types_list, localization_list)
        """
        df = pd.read_excel(self.excel_file, sheet_name='Complaint Type Master')

        complaint_types = []
        localizations = []
        current_parent = None
        localized_parent_types = set()

        # If no mapping provided, create empty dict
        if dept_name_to_code is None:
            dept_name_to_code = {}

        for _, row in df.iterrows():
            # Check if this is a parent row (has Complaint Type* filled)
            if pd.notna(row.get('Complaint Type*')):
                parent_type = str(row['Complaint Type*']).strip()

                # Get department name and convert to code
                dept_name = str(row['Department Name*']).strip() if pd.notna(row.get('Department Name*')) else None
                dept_code = dept_name_to_code.get(dept_name, dept_name) if dept_name else None

                current_parent = {
                    'type': parent_type,
                    'department': dept_code,
                    'slaHours': int(float(row['Resolution Time (Hours)*'])) if pd.notna(row.get('Resolution Time (Hours)*')) else None,
                    'keywords': str(row['Search Words (comma separated)*']).strip() if pd.notna(row.get('Search Words (comma separated)*')) else "",
                    'order': int(float(row['Priority'])) if pd.notna(row.get('Priority')) else None
                }

                # Auto-generate localization for parent type (only once)
                if parent_type not in localized_parent_types:
                    parent_type_code = ''.join(word.capitalize() for word in parent_type.split())
                    loc_code = f"SERVICEDFS.{parent_type_code.upper()}"
                    localizations.append({
                        'code': loc_code,
                        'message': parent_type,
                        'module': 'rainmaker-pgr',
                        'locale': 'en_IN'
                    })
                    localized_parent_types.add(parent_type)

            # Every row should have a sub-type
            if pd.notna(row.get('Complaint sub type*')) and str(row['Complaint sub type*']).strip() != '':
                sub_type_name = str(row['Complaint sub type*']).strip()

                # Auto-generate service code from sub-type name
                service_code = ''.join(word.capitalize() for word in sub_type_name.split())

                menu_path_value = current_parent['type'] if current_parent else sub_type_name

                ct = {
                    'serviceCode': service_code,
                    'name': sub_type_name,
                    'menuPath': menu_path_value,
                    'menuPathName': menu_path_value,
                    'active': True
                }

                # Add parent-level fields
                if current_parent:
                    if current_parent.get('department'):
                        ct['department'] = current_parent['department']
                    if current_parent.get('slaHours'):
                        ct['slaHours'] = current_parent['slaHours']
                    # Always include keywords, use empty string if not provided
                    ct['keywords'] = current_parent.get('keywords', "")
                    if current_parent.get('order'):
                        ct['order'] = current_parent['order']

                complaint_types.append(ct)

                # Auto-generate localization for sub-type
                loc_code = f"SERVICEDFS.{service_code.upper()}"
                localizations.append({
                    'code': loc_code,
                    'message': sub_type_name,
                    'module': 'rainmaker-pgr',
                    'locale': 'en_IN'
                })

        return complaint_types, localizations

    # ========================================================================
    # OTHER UTILITY READERS (Employee, Boundary, Workflow, Localization)
    # ========================================================================


    def read_employees_bulk(self, tenant_id: str, uploader=None):
        """Read employee data from simplified Employee Master sheet for bulk upload
        Converts department/designation/role NAMES to CODES internally

        Args:
            tenant_id: Target tenant ID (e.g., 'pg.citya')
            uploader: Authenticated APIUploader instance (required)

        Returns:
            list: List of employee objects ready for HRMS API
        """
        if uploader is None:
            raise ValueError("uploader parameter is required. Pass an authenticated APIUploader instance.")

        df = pd.read_excel(self.excel_file, sheet_name='Employee Master')

        # Fetch departments and create name->code mapping
        departments = uploader.fetch_departments(tenant_id)
        dept_name_to_code = {d.get('name'): d.get('code') for d in departments}

        # Fetch designations and create name->code mapping
        designations = uploader.fetch_designations(tenant_id)
        desig_name_to_code = {d.get('name'): d.get('code') for d in designations}

        # Fetch roles and create name->code mapping
        roles_list = uploader.fetch_roles(tenant_id)
        role_name_to_code = {r.get('name'): r.get('code') for r in roles_list}

        employees = []
        for idx, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get('User Name*')):
                continue

            user_name = str(row['User Name*']).strip()

            # Auto-generate employee code from user name
            # Remove special characters, convert to uppercase, replace spaces with underscore
            emp_code = ''.join(c if c.isalnum() or c.isspace() else '' for c in user_name)
            emp_code = emp_code.upper().replace(' ', '_')
            # If code is too long, truncate and add index
            if len(emp_code) > 20:
                emp_code = emp_code[:17] + f"_{idx}"

            mobile = str(row['Mobile Number*']).strip()

            # Get password (optional, defaults to eGov@123)
            password = str(row.get('Password', 'eGov@123')).strip() if pd.notna(row.get('Password')) else 'eGov@123'

            # Convert Excel dates to timestamps (milliseconds)
            def excel_date_to_timestamp(excel_date):
                """Convert Excel date to timestamp in milliseconds"""
                if pd.isna(excel_date):
                    return None
                # If already a timestamp (number > 1000000000000), return as-is
                if isinstance(excel_date, (int, float)) and excel_date > 1000000000000:
                    return int(excel_date)
                # If it's a pandas datetime, convert to timestamp
                if isinstance(excel_date, pd.Timestamp):
                    return int(excel_date.timestamp() * 1000)
                # If it's a string date, parse it
                if isinstance(excel_date, str):
                    from datetime import datetime
                    dt = pd.to_datetime(excel_date)
                    return int(dt.timestamp() * 1000)
                # Default: assume it's Excel serial date
                try:
                    dt = pd.to_datetime(excel_date, unit='D', origin='1899-12-30')
                    return int(dt.timestamp() * 1000)
                except:
                    return int(excel_date)

            # Parse dates from Excel
            from_date = excel_date_to_timestamp(row.get('Assignment From Date*'))
            if not from_date:
                from_date = 1725494400000  # Default

            appointment_date = excel_date_to_timestamp(row.get('Date of Appointment*'))
            if not appointment_date:
                appointment_date = 1718841600000  # Default

            # Convert department NAME to CODE
            dept_name = str(row.get('Department Name*', '')).strip()
            department = dept_name_to_code.get(dept_name, dept_name)  # Fallback to name if not found

            # Convert designation NAME to CODE
            desig_name = str(row.get('Designation Name*', '')).strip()
            designation = desig_name_to_code.get(desig_name, desig_name)  # Fallback to name if not found

            # Parse role NAMES and convert to CODES
            role_names_str = str(row.get('Role Names (comma separated)*', '')).strip()
            role_names = [r.strip() for r in role_names_str.split(',') if r.strip()]
            role_codes = [role_name_to_code.get(name, name) for name in role_names]  # Convert names to codes

            # Build roles list for both user and jurisdiction
            roles = []
            for i, role_code in enumerate(role_codes):
                # Get the original role name for this code
                role_name = role_names[i] if i < len(role_names) else role_code

                roles.append({
                    'code': role_code,
                    'name': role_name,
                    'tenantId': tenant_id
                })

            # Get boundary info (defaults to City level)
            hierarchy = str(row.get('Hierarchy Type', 'ADMIN')).strip()
            boundary_type = str(row.get('Boundary Type', 'City')).strip()
            boundary = str(row.get('Boundary Code', tenant_id.split('.')[0])).strip()

            # Build employee object
            emp = {
                'tenantId': tenant_id,
                'code': emp_code,
                'employeeStatus': str(row.get('Employee Status', 'EMPLOYED')).strip(),
                'employeeType': str(row.get('Employee Type', 'PERMANENT')).strip(),
                'dateOfAppointment': appointment_date,
                'assignments': [{
                    'fromDate': from_date,
                    'isCurrentAssignment': True,
                    'department': department,
                    'designation': designation
                }],
                'jurisdictions': [{
                    'hierarchy': hierarchy,
                    'boundaryType': boundary_type,
                    'boundary': boundary,
                    'tenantId': tenant_id,
                    'roles': roles  # Same roles assigned to jurisdiction
                }],
                'user': {
                    'name': user_name,
                    'mobileNumber': mobile,
                    'active': True,
                    'type': 'EMPLOYEE',
                    'tenantId': tenant_id,
                    'roles': roles,  # Same roles assigned to user
                    'password': password,  # Use password from Excel or default
                    'otpReference': '12345'
                },
                'serviceHistory': [],
                'education': [],
                'tests': []
            }

            # Add optional gender if provided
            if pd.notna(row.get('Gender')):
                emp['user']['gender'] = str(row['Gender']).strip()

            employees.append(emp)

        return employees

    def read_localization(self):
        """Read localization with auto-determination of module and locale based on code pattern"""
        try:
            df = pd.read_excel(self.excel_file, sheet_name='Localization')
        except:
            # Try lowercase sheet name
            try:
                df = pd.read_excel(self.excel_file, sheet_name='localization')
            except:
                return []

        if len(df) == 0:
            return []

        localizations = []

        for _, row in df.iterrows():
            # Skip rows with missing required fields
            if pd.notna(row.get('Code')) and pd.notna(row.get('Message')):
                code = str(row['Code']).strip()
                message = str(row['Message']).strip()

                # Skip if code or message is empty after stripping
                if not code or not message:
                    continue

                # Determine module and locale based on code pattern
                if code.startswith('SERVICEDFS.'):
                    # Service definitions → rainmaker-pgr
                    module = 'rainmaker-pgr'
                    locale = 'en_IN'
                elif code.startswith('COMMON_MASTERS_') or code.startswith('TENANT_TENANTS_'):
                    # Common masters (departments, designations, tenants) → rainmaker-common
                    module = 'rainmaker-common'
                    locale = 'en_IN'
                else:
                    # Default fallback
                    module = 'rainmaker-common'
                    locale = 'en_IN'

                localizations.append({
                    'code': code,
                    'message': message,
                    'module': module,
                    'locale': locale
                })

        return localizations


# ============================================================================
# WORKFLOW BUILDER
# ============================================================================



# ============================================================================
# API UPLOADER CLASS
# ============================================================================

class APIUploader:
    """Handles API uploads for PGR master data

    Service URLs are hardcoded and NOT configurable:
    - MDMS Service: :8094 (tenant, departments, designations, complaint types, employees)
    - Boundary Service: :8081 (boundaries, hierarchy, relationships)
    - Workflow Service: :8280 (workflow business services)
    - Localization Service: :8087 (localization/translations)
    """

    def __init__(self, base_url=None, username=None, password=None, user_type=None, tenant_id=None):
        """Initialize APIUploader with gateway authentication

        Args:
            base_url: Gateway URL (e.g., https://unified-dev.digit.org)
            username: Username for OAuth (e.g., DEV_SUPER_ADMIN)
            password: Password for OAuth
            user_type: EMPLOYEE or CITIZEN (default: EMPLOYEE)
            tenant_id: Tenant ID (e.g., dev, pg)
        """
        # Base gateway URL - same for all services (must be provided)
        if not base_url:
            raise ValueError("base_url is required. Please provide gateway URL.")

        self.base_url = base_url
        if self.base_url.endswith('/'):
            self.base_url = self.base_url[:-1]

        # Service endpoints from .env (configurable)
        # MDMS path varies by environment: /mdms-v2 (chakshu) or /egov-mdms-service (unified-dev)
        mdms_v2_service = os.getenv("MDMS_V2_SERVICE", None)  # Auto-detect if not set
        boundary_service = os.getenv("BOUNDARY_SERVICE", "/boundary-service")
        boundary_mgmt_service = os.getenv("BOUNDARY_MGMT_SERVICE", "/egov-bndry-mgmnt")
        localization_service = os.getenv("LOCALIZATION_SERVICE", "/localization")
        workflow_service = os.getenv("WORKFLOW_SERVICE", "/egov-workflow-v2")
        filestore_service = os.getenv("FILESTORE_SERVICE", "/filestore")
        hrms_service = os.getenv("HRMS_SERVICE", "/egov-hrms")
        data_handler_service = os.getenv("DATA_HANDLER_SERVICE", "/default-data-handler")
        auth_service = os.getenv("AUTH_SERVICE", "/user")

        # Build full service URLs
        # Auto-detect MDMS path if not explicitly set
        if mdms_v2_service:
            self.mdms_url = f"{self.base_url}{mdms_v2_service}"
        else:
            self.mdms_url = self._detect_mdms_path()
        self.boundary_url = f"{self.base_url}{boundary_service}"
        self.boundary_mgmt_url = f"{self.base_url}{boundary_mgmt_service}"
        self.localization_url = f"{self.base_url}{localization_service}"
        self.workflow_url = f"{self.base_url}{workflow_service}"
        self.filestore_url = f"{self.base_url}{filestore_service}"
        self.hrms_url = f"{self.base_url}{hrms_service}"
        self.Datahandlerurl = f"{self.base_url}{data_handler_service}"
        self.auth_url = f"{self.base_url}{auth_service}"

        # OAuth credentials
        self.username = username
        self.password = password
        self.user_type = user_type or "EMPLOYEE"
        self.tenant_id = tenant_id or "dev"

        # Auth tokens (populated by authenticate())
        self.auth_token = None
        self.user_info = None
        self.authenticated = False

        # Auto-authenticate if credentials provided
        if self.username and self.password:
            self.authenticate()

    def _detect_mdms_path(self):
        """Auto-detect the MDMS v2 service path by testing available endpoints.

        Different environments use different paths:
        - /mdms-v2 (chakshu-digit, newer deployments)
        - /egov-mdms-service (unified-dev, older deployments)

        Returns:
            str: The working MDMS URL (base_url + path)
        """
        # Paths to try, in order of preference
        paths_to_try = ["/mdms-v2", "/egov-mdms-service"]

        for path in paths_to_try:
            test_url = f"{self.base_url}{path}/v2/_search"
            try:
                # Simple probe request - doesn't need auth for search on some envs
                response = requests.post(
                    test_url,
                    json={
                        "MdmsCriteria": {"tenantId": "default", "limit": 1},
                        "RequestInfo": {"apiId": "Rainmaker"}
                    },
                    headers={"Content-Type": "application/json"},
                    timeout=5
                )

                # If not 404, this path exists (even if auth required, we get 401/403 not 404)
                if response.status_code != 404:
                    print(f"   MDMS path auto-detected: {path} (status: {response.status_code})")
                    return f"{self.base_url}{path}"
                else:
                    print(f"   MDMS path {path}: 404, trying next...")

            except requests.exceptions.RequestException as e:
                print(f"   MDMS path {path}: error {e}, trying next...")
                continue

        # Default fallback
        print(f"   MDMS path: using default /mdms-v2 (no path responded)")
        return f"{self.base_url}/mdms-v2"

    def authenticate(self):
        """Authenticate using OAuth2 password grant and fetch user info"""
        try:
            # OAuth2 token endpoint
            token_url = f"{self.auth_url}/oauth/token"

            # Standard eGov OAuth2 credentials (hardcoded)
            import base64
            client_id = "egov-user-client"
            client_secret = ""
            credentials = f"{client_id}:{client_secret}"
            encoded_credentials = base64.b64encode(credentials.encode()).decode()

            headers = {
                'Content-Type': 'application/x-www-form-urlencoded',
                'Authorization': f'Basic {encoded_credentials}'
            }

            data = {
                'username': self.username,
                'password': self.password,
                'userType': self.user_type,
                'tenantId': self.tenant_id,
                'scope': 'read',
                'grant_type': 'password'
            }

            response = requests.post(token_url, headers=headers, data=data, timeout=30)

            if response.status_code == 200:
                token_data = response.json()

                # Extract access_token as authToken
                self.auth_token = token_data.get('access_token')

                # Extract UserRequest as userInfo
                self.user_info = token_data.get('UserRequest', {})

                if self.auth_token and self.user_info:
                    self.authenticated = True
                    print(f"✅ Authentication successful!")
                    print(f"   User: {self.user_info.get('userName', 'Unknown')}")
                    print(f"   Name: {self.user_info.get('name', 'Unknown')}")
                    print(f"   Tenant: {self.user_info.get('tenantId', 'Unknown')}")
                    print(f"   Token: {self.auth_token[:20]}...")
                    return True
                else:
                    print(f"❌ Authentication response missing access_token or UserRequest")
                    print(f"Response: {response.text[:200]}")
                    return False
            else:
                print(f"❌ Authentication failed: {response.status_code}")
                print(f"Response: {response.text}")
                return False

        except Exception as e:
            print(f"❌ Authentication error: {str(e)}")
            return False

    def _extract_error_message(self, error_text: str) -> str:
        """Extract clean error message from API error response

        Args:
            error_text: Raw error response text (JSON)

        Returns:
            Cleaned error message string
        """
        try:
            # Try to parse as JSON
            error_json = json.loads(error_text)

            # Look for Errors array in response
            if isinstance(error_json, dict):
                errors = error_json.get('Errors', [])
                if errors and isinstance(errors, list) and len(errors) > 0:
                    first_error = errors[0]
                    if isinstance(first_error, dict):
                        # Extract message and description
                        message = first_error.get('message', '')
                        description = first_error.get('description', '')
                        code = first_error.get('code', '')

                        # Build clean error message
                        parts = []
                        if message:
                            parts.append(message)
                        if description:
                            parts.append(description)
                        if code and not message:  # Only add code if no message
                            parts.append(code)

                        if parts:
                            return ' - '.join(parts)

            # If no structured error found, return truncated original text
            return error_text[:200]

        except (json.JSONDecodeError, Exception):
            # If JSON parsing fails, return truncated original text
            return error_text[:200]

    def search_mdms_data(self, schema_code: str, tenant: str, unique_identifiers: List[str] = None,
                         limit: int = 100, offset: int = 0, include_inactive: bool = True) -> List[Dict]:
        """Generic function to search MDMS v2 data

        Args:
            schema_code: MDMS schema code (e.g., 'common-masters.Department', 'tenant.tenants')
            tenant: Tenant ID
            unique_identifiers: Optional list of unique identifiers to filter by
            limit: Max number of records (default: 100)
            offset: Pagination offset (default: 0)
            include_inactive: If False, filter out soft-deleted records (default: True)

        Returns:
            list: List of data objects retrieved (with 'isActive' field added from wrapper)
        """
        url = f"{self.mdms_url}/v2/_search"

        # Override userInfo tenantId to match the request tenant
        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant

        # Build criteria
        criteria = {
            "tenantId": tenant,
            "schemaCode": schema_code,
            "limit": limit,
            "offset": offset
        }

        # Add unique identifiers if provided
        if unique_identifiers:
            criteria["uniqueIdentifiers"] = unique_identifiers

        # Add isActive filter if only active records requested
        if not include_inactive:
            criteria["isActive"] = True

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": user_info_copy,
                "msgId": f"{int(time.time() * 1000)}|en_IN"
            },
            "MdmsCriteria": criteria
        }

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            data = response.json()

            # Extract data list from response
            # API returns: {"mdms": [{"id": "...", "data": {...}, "isActive": true/false, ...}]}
            mdms_records = data.get('mdms', [])

            # Include isActive status in each data object for caller to check
            data_list = []
            for record in mdms_records:
                record_data = record.get('data', {}).copy()
                # Add wrapper's isActive status to data object
                record_data['_isActive'] = record.get('isActive', True)
                record_data['_uniqueIdentifier'] = record.get('uniqueIdentifier')
                data_list.append(record_data)

            return data_list

        except requests.exceptions.HTTPError as e:
            print(f"HTTP Error during MDMS search for {schema_code}: {str(e)}")
            return []
        except Exception as e:
            print(f"Error during MDMS search for {schema_code}: {str(e)}")
            return []

    def create_mdms_data(self, schema_code: str, data_list: List[Dict], tenant: str, 
                        sheet_name: str = None, excel_file: str = None):
            """
            Upload MDMS data and write status directly into the uploaded Excel file
            
            Args:
                schema_code: MDMS schema code
                data_list: List of data objects to upload
                tenant: Tenant ID
                sheet_name: Excel sheet name to update with status
                excel_file: Path to the uploaded Excel file
            """
            url = f"{self.mdms_url}/v2/_create/{schema_code}"

            results = {
                'created': 0,
                'exists': 0,
                'failed': 0,
                'errors': []
            }

            print(f"\n[UPLOADING] {schema_code}")
            print(f"   Tenant: {tenant}")
            print(f"   Records: {len(data_list)}")
            print(f"   API URL: {url}")
            print("="*60)

            # Track row-by-row status for Excel update
            row_statuses = []

            for i, data_obj in enumerate(data_list, 1):
                unique_id = (
                    data_obj.get('code') or
                    data_obj.get('serviceCode') or
                    data_obj.get('userName') or
                    str(i)
                )

                payload = {
                    "RequestInfo": {
                        "apiId": "Rainmaker",
                        "authToken": self.auth_token,
                        "userInfo": self.user_info,
                        "msgId": "1695889012604|en_IN",
                        "plainAccessRequest": {}
                    },
                    "Mdms": {
                        "tenantId": tenant,
                        "schemaCode": schema_code,
                        "uniqueIdentifier": unique_id,
                        "data": data_obj,
                        "isActive": True
                    }
                }

                headers = {'Content-Type': 'application/json'}
                status = "SUCCESS"
                status_code = 200
                error_message = ""

                try:
                    response = requests.post(url, json=payload, headers=headers)
                    status_code = response.status_code
                    response.raise_for_status()
                    print(f"   [OK] [{i}/{len(data_list)}] {unique_id}")
                    results['created'] += 1

                except requests.exceptions.HTTPError as e:
                    # Get status code - response.status_code is the correct attribute
                    status_code = e.response.status_code if hasattr(e, 'response') and e.response is not None else 500
                    error_text = e.response.text if hasattr(e, 'response') and e.response is not None else str(e)

                    # Extract clean error message from API response
                    error_message = self._extract_error_message(error_text) if error_text else str(e)[:200]

                    # Fail fast on auth errors
                    if status_code == 401:
                        print(f"\n   ❌ AUTHORIZATION FAILED - Cannot create MDMS data")
                        print(f"   The endpoint /mdms-v2/v2/_create/{schema_code} requires authentication.")
                        print(f"   Ask admin to add it to EGOV_OPEN_ENDPOINTS_WHITELIST.")
                        results['failed'] = len(data_list)
                        results['errors'].append({'error': 'Authorization failed (401) - endpoint not in whitelist'})
                        return results

                    if 'already exists' in error_text.lower() or 'duplicate' in error_text.lower():
                        print(f"   [EXISTS] [{i}/{len(data_list)}] {unique_id} (HTTP {status_code})")
                        results['exists'] += 1
                        status = "EXISTS"
                    else:
                        print(f"   [FAILED] [{i}/{len(data_list)}] {unique_id} (HTTP {status_code})")
                        print(f"   ERROR: {error_message}")
                        results['failed'] += 1
                        results['errors'].append({'id': unique_id, 'error': error_message})
                        status = "FAILED"

                except Exception as e:
                    error_message = str(e)[:200]
                    status_code = 0
                    status = "FAILED"
                    print(f"   [ERROR] [{i}/{len(data_list)}] {unique_id} - {error_message[:100]}")
                    results['failed'] += 1
                    results['errors'].append({'id': unique_id, 'error': error_message})

                # Store status for this row
                row_statuses.append({
                    'row_index': i,  # 1-based index matching Excel data rows
                    'status': status,
                    'status_code': status_code,
                    'error_message': error_message
                })

                time.sleep(0.1)

            # Summary
            print("="*60)
            print(f"[SUMMARY] Created: {results['created']}")
            print(f"[SUMMARY] Already Exists: {results['exists']}")
            print(f"[SUMMARY] Failed: {results['failed']}")
            print("="*60)

            # Write status columns directly into the uploaded Excel file
            if excel_file and sheet_name and row_statuses:
                self._write_status_to_excel(
                    excel_file=excel_file,
                    sheet_name=sheet_name,
                    row_statuses=row_statuses,
                    schema_code=schema_code
                )

            return results

    def delete_mdms_data(self, schema_code: str, tenant: str, unique_ids: List[str] = None) -> Dict:
        """Soft-delete MDMS data by setting isActive=false

        Args:
            schema_code: Schema code (e.g., 'common-masters.Department')
            tenant: Tenant ID
            unique_ids: Optional list of specific uniqueIdentifiers to delete.
                       If None, deletes ALL data for this schema/tenant.

        Returns:
            dict: {deleted: count, failed: count, errors: []}
        """
        print(f"\n🗑️ Deleting MDMS data: {schema_code} for tenant: {tenant}")

        results = {'deleted': 0, 'failed': 0, 'skipped': 0, 'errors': []}

        # Step 1: Search for existing data
        search_url = f"{self.mdms_url}/v2/_search"
        update_url = f"{self.mdms_url}/v2/_update/{schema_code}"

        search_payload = {
            "RequestInfo": {
                "apiId": "asset-services",
                "msgId": f"search-{int(time.time()*1000)}",
                "authToken": self.auth_token,
                "userInfo": self.user_info
            },
            "MdmsCriteria": {
                "tenantId": tenant,
                "schemaCode": schema_code,
                "limit": 500,
                "offset": 0
            }
        }

        try:
            response = requests.post(search_url, json=search_payload, headers={'Content-Type': 'application/json'})
            if response.status_code != 200:
                print(f"   ❌ Failed to search MDMS data: {response.status_code}")
                return results

            data = response.json()
            mdms_records = data.get('mdms', [])

            if not mdms_records:
                print(f"   ℹ️ No MDMS data found for {schema_code}")
                return results

            print(f"   Found {len(mdms_records)} records")

            # Filter by unique_ids if provided
            if unique_ids:
                mdms_records = [r for r in mdms_records if r.get('uniqueIdentifier') in unique_ids]
                print(f"   Filtered to {len(mdms_records)} records matching provided IDs")

            # Step 2: Update each record with isActive=false
            auth_failed = False
            for i, record in enumerate(mdms_records):
                unique_id = record.get('uniqueIdentifier', 'unknown')

                # Skip if already inactive
                if not record.get('isActive', True):
                    print(f"   ⏭️ Already inactive: {unique_id}")
                    results['skipped'] += 1
                    continue

                update_payload = {
                    "RequestInfo": {
                        "apiId": "asset-services",
                        "msgId": f"delete-{int(time.time()*1000)}",
                        "authToken": self.auth_token,
                        "userInfo": self.user_info
                    },
                    "Mdms": {
                        "tenantId": tenant,
                        "schemaCode": schema_code,
                        "uniqueIdentifier": unique_id,
                        "id": record.get('id'),
                        "data": record.get('data', {}),
                        "auditDetails": record.get('auditDetails'),
                        "isActive": False  # This is the key - set to false to "delete"
                    }
                }

                try:
                    upd_response = requests.post(update_url, json=update_payload, headers={'Content-Type': 'application/json'})
                    if upd_response.status_code == 200:
                        print(f"   ✅ Deleted: {unique_id}")
                        results['deleted'] += 1
                    elif upd_response.status_code == 401:
                        # Fail fast on auth errors
                        print(f"\n   ❌ AUTHORIZATION FAILED - Cannot delete MDMS data")
                        print(f"   The endpoint /mdms-v2/v2/_update/{schema_code} requires authentication.")
                        print(f"   Ask admin to add it to EGOV_OPEN_ENDPOINTS_WHITELIST or use direct DB access.")
                        results['failed'] = len(mdms_records) - results['skipped']
                        results['errors'].append({'error': 'Authorization failed (401) - endpoint not in whitelist'})
                        auth_failed = True
                        break
                    else:
                        error_msg = upd_response.text[:100]
                        print(f"   ❌ Failed to delete {unique_id}: {error_msg}")
                        results['failed'] += 1
                        results['errors'].append({'id': unique_id, 'error': error_msg})
                except Exception as e:
                    print(f"   ❌ Error deleting {unique_id}: {str(e)[:50]}")
                    results['failed'] += 1
                    results['errors'].append({'id': unique_id, 'error': str(e)})

                time.sleep(0.05)  # Small delay between updates

            if auth_failed:
                return results

        except Exception as e:
            print(f"   ❌ Error: {str(e)[:100]}")
            results['errors'].append({'error': str(e)})

        print(f"\n   Summary: Deleted {results['deleted']}, Failed {results['failed']}, Skipped {results['skipped']}")
        return results

    def rollback_mdms_by_schema(self, schema_codes: List[str], tenant: str) -> Dict:
        """Rollback (delete) all MDMS data for multiple schema codes

        Args:
            schema_codes: List of schema codes to rollback
            tenant: Tenant ID

        Returns:
            dict: {schema_code: {deleted, failed, ...}, ...}
        """
        print(f"\n{'='*60}")
        print(f"MDMS ROLLBACK")
        print(f"{'='*60}")
        print(f"Tenant: {tenant}")
        print(f"Schemas: {', '.join(schema_codes)}")

        results = {}
        for schema_code in schema_codes:
            results[schema_code] = self.delete_mdms_data(schema_code, tenant)

        # Summary
        total_deleted = sum(r.get('deleted', 0) for r in results.values())
        total_failed = sum(r.get('failed', 0) for r in results.values())

        print(f"\n{'─'*40}")
        print(f"Rollback Complete: {total_deleted} deleted, {total_failed} failed")
        print(f"{'─'*40}")

        return results

    def _write_status_to_excel(self, excel_file: str, sheet_name: str,
                               row_statuses: List[Dict], schema_code: str):
        """
        Write / overwrite _STATUS, _STATUS_CODE, _ERROR_MESSAGE columns directly into uploaded Excel.
        - If columns exist: overwrite in-place.
        - If columns do not exist: create exactly one set of new columns at the right-most side.
        - Does NOT use ws.append() (so it won't accidentally shift or insert rows).
        Note: row_statuses[*]['row_index'] is expected to be the Excel row number you want to write to.
        If your row_index is zero-based data-row index (0..n-1) change excel_row = header_row + 1 + row_index below.
        """
        try:
            print(f"\n📝 Updating Excel file: {excel_file}")
            print(f"   Sheet: {sheet_name}")
            
            wb = load_workbook(excel_file, data_only=False)
            if sheet_name not in wb.sheetnames:
                print(f"   ⚠️  Sheet '{sheet_name}' not found - skipping status update")
                return
            ws = wb[sheet_name]
    
            # --- Find header row and map existing headers (assume header in row 1) ---
            header_row = 1
            header_map = {}
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=header_row, column=col).value
                if isinstance(value, str) and value.strip():
                    header_map[value.strip()] = col
    
            # --- Determine/create columns in a safe, non-overlapping way ---
            max_col = ws.max_column
    
            # We'll create new columns only if header missing; if we create one column,
            # we update max_col so the next created column goes to max_col+1, etc.
            def get_or_create_col(header_name):
                nonlocal max_col
                if header_name in header_map:
                    return header_map[header_name]
                else:
                    # create new column at the right
                    max_col += 1
                    header_map[header_name] = max_col
                    # set header cell (style later)
                    ws.cell(row=header_row, column=max_col, value=header_name)
                    return max_col
    
            status_col = get_or_create_col("_STATUS")
            code_col = get_or_create_col("_STATUS_CODE")
            error_col = get_or_create_col("_ERROR_MESSAGE")
    
            # --- Styles ---
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True)
            success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            exists_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
            # Apply header styles (for newly created headers and to ensure consistent style)
            for header, col_idx in [("_STATUS", status_col), ("_STATUS_CODE", code_col), ("_ERROR_MESSAGE", error_col)]:
                hdr_cell = ws.cell(row=header_row, column=col_idx)
                hdr_cell.fill = header_fill
                hdr_cell.font = header_font
    
            # --- Write each row_status in-place (overwrite) ---
            for row_status in row_statuses:
                # === IMPORTANT: row_index is 1-based data row index (1, 2, 3, ...) ===
                # We need to convert it to actual Excel row number by adding header_row
                # For example: row_index=1 (first data row) → Excel row 2 (header is row 1)
                row_index = row_status['row_index']

                # Validate row_index is integer
                try:
                    row_index = int(row_index)
                except Exception:
                    # skip invalid rows
                    print(f"   ⚠️  Skipping invalid row_index: {row_status.get('row_index')}")
                    continue

                # Calculate actual Excel row: header_row + row_index
                excel_row = header_row + row_index

                status = row_status.get('status', '')
                status_code = row_status.get('status_code', '')
                error_msg = row_status.get('error_message', '')

                # Overwrite exact cells (openpyxl will create cell objects if row > current max)
                status_cell = ws.cell(row=excel_row, column=status_col, value=status)
                if status == "SUCCESS":
                    status_cell.fill = success_fill
                elif status == "EXISTS":
                    status_cell.fill = exists_fill
                elif status == "FAILED":
                    status_cell.fill = failed_fill
                else:
                    status_cell.fill = PatternFill(fill_type=None)

                ws.cell(row=excel_row, column=code_col, value=status_code)
                ws.cell(row=excel_row, column=error_col, value=error_msg)

            # --- Column widths only for newly added columns (or enforce widths always) ---
            # If you prefer to always set widths, remove the conditional checks.
            for col_idx, width in [(status_col, 15), (code_col, 15), (error_col, 50)]:
                ws.column_dimensions[get_column_letter(col_idx)].width = width
    
            # --- Protect status columns (lock cells) ---
            for r in range(1, ws.max_row + 1):
                for c in (status_col, code_col, error_col):
                    cell = ws.cell(row=r, column=c)
                    cell.protection = cell.protection.copy(locked=True)
    
            # Save
            wb.save(excel_file)
            wb.close()
            print(f"   ✅ Status columns updated successfully!")
            print(f"   📊 Updated {len(row_statuses)} rows")
    
        except Exception as e:
            print(f"   ⚠️  Could not update Excel: {str(e)}")


    def _generate_error_excel(self, failed_records: List[Dict], schema_code: str, sheet_name: str, dept_code_to_name: Dict = None) -> str:
        """Append failed records to a single consolidated error Excel file

        Args:
            failed_records: List of failed data records with status info
            schema_code: Schema code for naming
            sheet_name: Sheet name for the error file
            dept_code_to_name: Reverse mapping from department codes to names

        Returns:
            str: Path to the error Excel file
        """
        try:
            import pandas as pd
            from datetime import datetime
            import os
            from openpyxl import load_workbook

            # Create errors directory if it doesn't exist
            error_dir = 'errors'
            os.makedirs(error_dir, exist_ok=True)

            # Use a single consolidated filename
            filename = f"{error_dir}/FAILED_RECORDS.xlsx"

            # Transform records back to Excel template format
            transformed_records = self._transform_to_excel_format(failed_records, schema_code, dept_code_to_name)

            # Flatten nested structures into readable columns
            flattened_records = []
            for record in transformed_records:
                flat_record = {}

                for key, value in record.items():
                    # Handle nested objects (like city, jurisdiction, etc.)
                    if isinstance(value, dict):
                        # Flatten nested dict with prefix
                        for nested_key, nested_value in value.items():
                            flat_record[f"{key}.{nested_key}"] = nested_value
                    # Handle lists (like tenants in citymodule, roles, etc.)
                    elif isinstance(value, list):
                        if len(value) > 0:
                            # For list of dicts (like tenants: [{'code': 'pg'}, {'code': 'pg.citya'}])
                            if isinstance(value[0], dict):
                                # Extract codes/names and join with comma
                                codes = [item.get('code', item.get('name', str(item))) for item in value]
                                flat_record[key] = ', '.join(str(c) for c in codes)
                            else:
                                # Simple list (like pincodes)
                                flat_record[key] = ', '.join(str(v) for v in value)
                        else:
                            flat_record[key] = ''
                    else:
                        flat_record[key] = value

                flattened_records.append(flat_record)

            # Convert to DataFrame
            df = pd.DataFrame(flattened_records)

            # Define columns to exclude (internal/auto-generated fields only)
            exclude_cols = [
                'active', 'isActive', 'tenantId', 'uniqueIdentifier'
            ]

            # Remove excluded columns
            cols_to_keep = [col for col in df.columns if col not in exclude_cols]
            df = df[cols_to_keep]

            # Reorder columns to put status columns at the END
            status_cols = ['_STATUS', '_STATUS_CODE', '_ERROR_MESSAGE']
            other_cols = [col for col in df.columns if col not in status_cols]
            df = df[other_cols + status_cols]

            # Check if file exists
            if os.path.exists(filename):
                # Append to existing file
                with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # Create new file
                df.to_excel(filename, sheet_name=sheet_name, index=False)

            # Apply sheet protection to error columns
            self._protect_error_columns(filename, sheet_name, len(other_cols))

            print(f"\n📊 ERROR REPORT UPDATED:")
            print(f"   File: {filename}")
            print(f"   Sheet: {sheet_name}")
            print(f"   Failed Records: {len(failed_records)}")
            print(f"   💡 Fix the errors and re-upload this file (Error columns are protected)")

            return filename

        except Exception as e:
            print(f"\n⚠️  Could not generate error Excel: {str(e)}")
            return None

    def _transform_to_excel_format(self, records: List[Dict], schema_code: str, dept_code_to_name: Dict = None) -> List[Dict]:
        """Transform API payload format back to Excel template format

        Args:
            records: List of records in API format
            schema_code: Schema code to determine transformation rules
            dept_code_to_name: Reverse mapping from department codes to names

        Returns:
            List of records in Excel template format
        """
        if dept_code_to_name is None:
            dept_code_to_name = {}

        transformed = []

        # Handle Departments - rename 'code' and 'name' to match template
        if 'Department' in schema_code:
            for record in records:
                excel_record = {
                    'Department Name*': record.get('name', ''),
                    '_STATUS': record.get('_STATUS'),
                    '_STATUS_CODE': record.get('_STATUS_CODE'),
                    '_ERROR_MESSAGE': record.get('_ERROR_MESSAGE')
                }
                transformed.append(excel_record)

        # Handle Designations - show department name instead of code
        elif 'Designation' in schema_code:
            for record in records:
                dept_code = record.get('departmentCode', '')
                dept_name = dept_code_to_name.get(dept_code, dept_code)

                excel_record = {
                    'Department Name*': dept_name,
                    'Designation Name*': record.get('name', ''),
                    '_STATUS': record.get('_STATUS'),
                    '_STATUS_CODE': record.get('_STATUS_CODE'),
                    '_ERROR_MESSAGE': record.get('_ERROR_MESSAGE')
                }
                transformed.append(excel_record)

        # Handle Complaint Types - show department name and extract complaint type/subtype
        elif 'ServiceDefs' in schema_code:
            for record in records:
                dept_code = record.get('department', '')
                dept_name = dept_code_to_name.get(dept_code, dept_code)

                excel_record = {
                    'Complaint Type*': record.get('menuPath', ''),
                    'Complaint sub type*': record.get('name', ''),
                    'Department Name*': dept_name,
                    'Resolution Time (Hours)*': record.get('slaHours', ''),
                    'Search Words (comma separated)*': record.get('keywords', ''),
                    'Priority': record.get('priority', ''),
                    '_STATUS': record.get('_STATUS'),
                    '_STATUS_CODE': record.get('_STATUS_CODE'),
                    '_ERROR_MESSAGE': record.get('_ERROR_MESSAGE')
                }
                transformed.append(excel_record)

        # Default: return records as-is (ensuring status fields are preserved)
        else:
            # Make sure status fields are always included
            for record in records:
                excel_record = record.copy()
                # Ensure status fields exist
                if '_STATUS' not in excel_record:
                    excel_record['_STATUS'] = record.get('_STATUS')
                if '_STATUS_CODE' not in excel_record:
                    excel_record['_STATUS_CODE'] = record.get('_STATUS_CODE')
                if '_ERROR_MESSAGE' not in excel_record:
                    excel_record['_ERROR_MESSAGE'] = record.get('_ERROR_MESSAGE')
                transformed.append(excel_record)

        return transformed

    def _protect_error_columns(self, filename: str, sheet_name: str, data_col_count: int):
        """Protect the error status columns (last 3 columns) in the Excel sheet

        Args:
            filename: Path to Excel file
            sheet_name: Sheet name to protect
            data_col_count: Number of data columns (non-error columns)
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter

            wb = load_workbook(filename)
            ws = wb[sheet_name]

            # Get total columns
            total_cols = ws.max_column

            # Error columns are the last 3 columns
            error_col_start = data_col_count + 1  # +1 because Excel is 1-indexed

            # Apply gray background and bold font to error column headers
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            bold_font = Font(bold=True)

            for col_idx in range(error_col_start, total_cols + 1):
                col_letter = get_column_letter(col_idx)

                # Style header
                header_cell = ws[f'{col_letter}1']
                header_cell.fill = gray_fill
                header_cell.font = bold_font

                # Lock all cells in error columns (data + header)
                for row in range(1, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    cell.protection = cell.protection.copy(locked=True)

            # Unlock all data columns so users can edit them
            for col_idx in range(1, error_col_start):
                col_letter = get_column_letter(col_idx)
                for row in range(2, ws.max_row + 1):  # Skip header row
                    cell = ws[f'{col_letter}{row}']
                    cell.protection = cell.protection.copy(locked=False)

            # Protect the sheet (allow users to edit unlocked cells only)
            ws.protection.sheet = True
            ws.protection.password = ''  # Empty password (no password)
            ws.protection.selectLockedCells = True
            ws.protection.selectUnlockedCells = True

            wb.save(filename)

        except Exception as e:
            # Don't fail if protection doesn't work
            print(f"   ⚠️  Could not apply sheet protection: {str(e)}")

    def create_localization_messages(self, localization_list: List[Dict], tenant: str, sheet_name: str = 'Localization'):
        """Upload localization messages via localization service API"""
        url = f"{self.localization_url}/messages/v1/_upsert"

        results = {
            'created': 0,
            'exists': 0,
            'failed': 0,
            'errors': [],
            'failed_records': []
        }

        print(f"\n[UPLOADING] Localization Messages")
        print(f"   Tenant: {tenant}")
        print(f"   Total Messages: {len(localization_list)}")
        print(f"   API URL: {url}")
        print("="*60)

        # Group messages by locale for batch upload
        from collections import defaultdict
        by_locale = defaultdict(list)
        for loc in localization_list:
            by_locale[loc['locale']].append(loc)

        print(f"\n   Found {len(by_locale)} locales: {', '.join(by_locale.keys())}")
        print("="*60)

        # Upload each locale batch - split into smaller chunks to avoid 413 error
        BATCH_SIZE = 500  # Upload 500 messages at a time

        for locale, messages in by_locale.items():
            total_messages = len(messages)
            print(f"   📤 Locale: {locale} - Uploading {total_messages} messages in batches of {BATCH_SIZE}...")

            # Split into batches
            for batch_idx in range(0, total_messages, BATCH_SIZE):
                batch = messages[batch_idx:batch_idx + BATCH_SIZE]
                batch_num = (batch_idx // BATCH_SIZE) + 1
                total_batches = (total_messages + BATCH_SIZE - 1) // BATCH_SIZE

                payload = {
                    "RequestInfo": {
                        "apiId": "emp",
                        "ver": "1.0",
                        "action": "create",
                        "msgId": f"{int(time.time() * 1000)}",
                        "authToken": self.auth_token,
                        "userInfo": self.user_info
                    },
                    "locale": locale,
                    "tenantId": tenant,
                    "messages": batch
                }

                headers = {'Content-Type': 'application/json'}
                status_code = None

                try:
                    response = requests.post(url, json=payload, headers=headers, timeout=120)
                    status_code = response.status_code
                    response.raise_for_status()
                    print(f"      ✅ Batch {batch_num}/{total_batches}: {len(batch)} messages uploaded")
                    results['created'] += len(batch)

                except requests.exceptions.HTTPError as e:
                    status_code = e.response.status_code if hasattr(e, 'response') and e.response is not None else 500
                    error_text = e.response.text if hasattr(e, 'response') and e.response is not None else str(e)

                    # Extract clean error message from API response
                    error_message = self._extract_error_message(error_text) if error_text else str(e)[:200]

                    # Fail fast on auth errors
                    if status_code == 401:
                        print(f"\n   ❌ AUTHORIZATION FAILED - Cannot upload localization messages")
                        print(f"   The endpoint /localization/messages/v1/_upsert requires authentication.")
                        print(f"   Ask admin to add it to EGOV_OPEN_ENDPOINTS_WHITELIST.")
                        results['failed'] = total_messages
                        results['errors'].append({'error': 'Authorization failed (401) - endpoint not in whitelist'})
                        return results

                    # Check for duplicate/already exists errors
                    if ('duplicate' in error_text.lower() or
                        'already exists' in error_text.lower() or
                        'DUPLICATE_RECORDS' in error_text or
                        'DuplicateMessageIdentityException' in error_text or
                        'unique_message_entry' in error_text.lower()):
                        print(f"      ⚠️ Batch {batch_num}/{total_batches}: {len(batch)} messages already exist")
                        results['exists'] += len(batch)
                        # DON'T add to failed_records - already exists is not a failure!
                    else:
                        # True failure
                        print(f"      ❌ Batch {batch_num}/{total_batches} FAILED (HTTP {status_code})")
                        print(f"         ERROR: {error_message}")
                        results['failed'] += len(batch)
                        results['errors'].append({
                            'locale': locale,
                            'batch': batch_num,
                            'count': len(batch),
                            'error': error_message
                        })

                        # Store failed messages for Excel export - ONLY TRUE FAILURES
                        for msg in batch:
                            failed_record = msg.copy()
                            failed_record['_STATUS'] = 'FAILED'
                            failed_record['_STATUS_CODE'] = status_code
                            failed_record['_ERROR_MESSAGE'] = error_message
                            results['failed_records'].append(failed_record)

                except Exception as e:
                    error_message = str(e)[:200]
                    status_code = 0
                    print(f"      ❌ Batch {batch_num}/{total_batches} ERROR: {error_message}")
                    results['failed'] += len(batch)
                    results['errors'].append({
                        'locale': locale,
                        'batch': batch_num,
                        'count': len(batch),
                        'error': error_message
                    })

                    # Store failed messages for Excel export
                    for msg in batch:
                        failed_record = msg.copy()
                        failed_record['_STATUS'] = 'FAILED'
                        failed_record['_STATUS_CODE'] = status_code
                        failed_record['_ERROR_MESSAGE'] = error_message
                        results['failed_records'].append(failed_record)

            time.sleep(0.2)

        # Summary
        print("="*60)
        print(f"[SUMMARY] Created: {results['created']}")
        print(f"[SUMMARY] Already Exists: {results['exists']}")
        print(f"[SUMMARY] Failed: {results['failed']}")

        if results['errors']:
            print(f"\n[ERRORS] Found {len(results['errors'])} error(s):")
            for err in results['errors']:
                print(f"   - Locale: {err['locale']} ({err['count']} messages)")
                print(f"     Error: {err['error'][:100]}")

        # Generate error Excel if there are failures
        if results['failed_records'] and sheet_name:
            error_file = self._generate_error_excel(results['failed_records'], 'localization.messages', sheet_name)
            results['error_file'] = error_file

        print("="*60)

        return results

    def create_boundary_hierarchy(self, hierarchy_data: Dict):
        """Create boundary hierarchy definition

        Args:
            hierarchy_data: Dict containing tenantId, hierarchyType, and boundaryHierarchy

        Returns:
            API response dict
        """
        url = f"{self.boundary_url}/boundary-hierarchy-definition/_create"

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": self.user_info,
                "msgId": "1695889012604|en_IN",
                "plainAccessRequest": {}
            },
            'BoundaryHierarchy': hierarchy_data
        }
     

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
          
            response.raise_for_status()
            print(f"\n✅ [SUCCESS] Boundary hierarchy created")
            print(f"   Tenant: {hierarchy_data.get('tenantId')}")
            print(f"   Hierarchy Type: {hierarchy_data.get('hierarchyType')}")
            print(f"   Levels: {len(hierarchy_data.get('boundaryHierarchy', []))}")
            return response.json()
        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)

            # Check if hierarchy already exists
            if 'already exists' in error_text.lower() or 'duplicate' in error_text.lower():
                print(f"\n⚠️  [EXISTS] Boundary hierarchy already exists")
                print(f"   Tenant: {hierarchy_data.get('tenantId')}")
                print(f"   Hierarchy Type: {hierarchy_data.get('hierarchyType')}")
                return {'success': True, 'exists': True}
            else:
                print(f"\n❌ [ERROR] Failed: HTTP {e.response.status_code}")
                print(f"   ERROR Details: {error_text[:500]}")
                raise
        except Exception as e:
            print(f"\n❌ [ERROR] Failed: {str(e)}")
            raise

    def update_stateinfo_language(self, language_label: str, language_value: str, state_tenant: str = "pg"):
        """
        Update common-masters.StateInfo to add a new language to the state-level configuration

        Args:
            language_label: Display name of language (e.g., 'हिंदी', 'ਪੰਜਾਬੀ')
            language_value: Locale code (e.g., 'hi_IN', 'pa_IN')
            state_tenant: State tenant ID (default: 'pg')

        Returns:
            Dict with results of update operation
        """
        search_url = f"{self.mdms_url}/v2/_search"
        update_url = f"{self.mdms_url}/v2/_update/common-masters.StateInfo"

        result = {
            'updated': False,
            'error': None
        }

        print(f"\n[UPDATING STATE INFO LANGUAGES]")
        print(f"   Adding: {language_label} ({language_value})")
        print(f"   State Tenant: {state_tenant}")
        print("="*60)

        try:
            # Step 1: Search for existing StateInfo data (without uniqueIdentifier filter)
            user_info_copy = self.user_info.copy()
            user_info_copy['tenantId'] = state_tenant

            search_payload = {
                "RequestInfo": {
                    "apiId": "Rainmaker",
                    "authToken": self.auth_token,
                    "userInfo": user_info_copy,
                    "msgId": f"{int(time.time() * 1000)}|en_IN"
                },
                "MdmsCriteria": {
                    "tenantId": state_tenant,
                    "schemaCode": "common-masters.StateInfo"
                    # Don't filter by uniqueIdentifiers - get all StateInfo records
                }
            }

            headers = {'Content-Type': 'application/json'}
            search_response = requests.post(search_url, json=search_payload, headers=headers)
            search_response.raise_for_status()
            search_data = search_response.json()

            # Check if StateInfo exists
            if not search_data.get('mdms') or len(search_data['mdms']) == 0:
                print(f"   [SKIP] StateInfo not found for tenant {state_tenant}")
                result['error'] = 'StateInfo not found in database'
                return result

            # Get the first StateInfo record (usually there's only one per state)
            mdms_record = search_data['mdms'][0]
    

            mdms_record = search_data['mdms'][0]
            stateinfo_data = mdms_record['data']

            # Step 2: Check if language already exists
            existing_languages = stateinfo_data.get('languages', [])
            language_exists = any(
                lang.get('value') == language_value
                for lang in existing_languages
            )

            if language_exists:
                print(f"   [EXISTS] Language already exists in StateInfo")
                result['updated'] = False
                return result

            # Step 3: Add new language to the list
            new_language = {
                "label": language_label,
                "value": language_value
            }
            stateinfo_data['languages'] = existing_languages + [new_language]

            # Step 4: Update StateInfo with new language
            user_info_update = self.user_info.copy()
            user_info_update['tenantId'] = state_tenant

            # Get uniqueIdentifier from the existing record or use 'code' field
            unique_id = mdms_record.get('uniqueIdentifier') or stateinfo_data.get('code', state_tenant)
            print(f"   Using uniqueIdentifier: {unique_id}")

            # Clean null values
            def clean_nulls(obj):
                if isinstance(obj, dict):
                    return {k: clean_nulls(v) for k, v in obj.items() if v is not None}
                elif isinstance(obj, list):
                    return [clean_nulls(item) for item in obj if item is not None]
                else:
                    return obj

            stateinfo_data = clean_nulls(stateinfo_data)

            update_payload = {
                "RequestInfo": {
                    "apiId": "asset-services",
                    "msgId": "update-stateinfo-language",
                    "authToken": self.auth_token,
                    "userInfo": user_info_update
                },
                "Mdms": {
                    "tenantId": state_tenant,
                    "schemaCode": "common-masters.StateInfo",
                    "uniqueIdentifier": unique_id,
                    "id": mdms_record['id'],
                    "data": stateinfo_data,
                    "auditDetails": mdms_record.get('auditDetails'),
                    "isActive": mdms_record.get('isActive', True)
                }
            }

            update_response = requests.post(update_url, json=update_payload, headers=headers)
            update_response.raise_for_status()

            print(f"   [OK] StateInfo updated successfully with new language")
            result['updated'] = True

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)
            print(f"   [FAILED] HTTP Error: {error_text[:200]}")
            result['error'] = error_text[:200]

        except Exception as e:
            print(f"   [ERROR] {str(e)[:100]}")
            result['error'] = str(e)[:200]

        print("="*60)
        return result

    def setup_default_data(self, targetTenantId: str, module: str, schemaCodes: list, onlySchemas: bool = False) -> dict:

        url = f"{self.Datahandlerurl}/tenant/new"

        payload = {
            "RequestInfo": {
                "apiId": "default-data-handler",
                "ver": "1.0",
                "ts": None,
                "action": "create",
                "msgId": "default-data-setup",
                "authToken": self.auth_token,
                "userInfo": self.user_info
            },
            "targetTenantId": targetTenantId,
            "module": module,
            "schemaCodes": schemaCodes,
            "onlySchemas": onlySchemas
        }

        print("\n[DEFAULT DATA SETUP]")
        print(f"   Target Tenant: {targetTenantId}")
        print(f"   Module:        {module}")
        print(f"   Schemas:       {schemaCodes}")
        print(f"   Only Schemas:  {onlySchemas}")
        print(f"   API URL:       {url}")
        print("="*60)

        try:
            # Tenant creation can take 5-10 minutes, use long timeout
            print(f"   ⏳ This may take 5-10 minutes, please wait...")
            response = requests.post(url, json=payload, headers={"Content-Type": "application/json"}, timeout=1200)
            response.raise_for_status()

            result = response.json()

            print(f"   [SUCCESS] Default data setup complete for {targetTenantId}")
            print("Once the new tenant is created, please log in again using the new root tenant admin credentials.")
            print("="*60)

            return {
                "success": True,
                "response": result
            }

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, "text") else str(e)
            print(f"[FAILED] HTTP ERROR {e.response.status_code}")
            print(error_text)
            print("="*60)
            return {"success": False, "error": error_text}

        except Exception as e:
            print(f"[ERROR] {str(e)}")
            print("="*60)
            return {"success": False, "error": str(e)}

    # ========================================================================
    # BOUNDARY MANAGEMENT METHODS
    # ========================================================================

    def search_boundary_hierarchies(self, tenant_id: str, limit: int = 100, offset: int = 0) -> List[Dict]:
        """Search for existing boundary hierarchies

        Args:
            tenant_id: Tenant ID to search for
            limit: Maximum number of results
            offset: Offset for pagination

        Returns:
            List of boundary hierarchy definitions
        """
        url = f"{self.boundary_url}/boundary-hierarchy-definition/_search"

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": self.user_info,
                "msgId": "1695889012604|en_IN",
                "plainAccessRequest": {}
            },
            "BoundaryTypeHierarchySearchCriteria": {
                "tenantId": tenant_id,
                "limit": limit,
                "offset": offset
            }
        }
        

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            data = response.json()

        

            hierarchies = data.get('BoundaryHierarchy', [])

            if hierarchies is None:
                print(f"\n⚠️ BoundaryHierarchy returned None (no hierarchies exist for {tenant_id})")
                return []

            print(f"\n✅ Found {len(hierarchies)} boundary hierarchy(ies) for tenant: {tenant_id}")

            for h in hierarchies:
                print(f"   • {h.get('hierarchyType', 'UNKNOWN')} ({len(h.get('boundaryHierarchy', []))} levels)")

            return hierarchies

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)
            print(f"❌ HTTP Error: {error_text[:200]}")
            return []
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            import traceback
            print(f"📋 Traceback: {traceback.format_exc()[:300]}")
            return []

    def generate_boundary_template(self, tenant_id: str, hierarchy_type: str, force_update: bool = True) -> Dict:
        """Generate boundary template file

        Args:
            tenant_id: Tenant ID
            hierarchy_type: Hierarchy type (e.g., 'ADMIN', 'REVENUE')
            force_update: Force regeneration

        Returns:
            Dict with generation task details
        """
        url = f"{self.boundary_mgmt_url}/v1/_generate"

        params = {
            "tenantId": tenant_id,
            "hierarchyType": hierarchy_type,
            "forceUpdate": str(force_update).lower()
        }

        # Override userInfo tenantId to match the request tenant
        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant_id

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": user_info_copy,
                "msgId": f"{int(time.time() * 1000)}|en_IN",
                "plainAccessRequest": {}
            }
        }
       

        headers = {'Content-Type': 'application/json'}

        try:
            response = requests.post(url, json=payload, headers=headers, params=params)
            response.raise_for_status()
            data = response.json()

            resource = data.get('ResourceDetails', [{}])[0]
            print(f"\n✅ Template generation initiated")
            print(f"   Task ID: {resource.get('id')}")
            print(f"   Status: {resource.get('status')}")

            return resource

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e.response, 'text') else str(e)
            print(f"❌ HTTP Error: {error_text[:300]}")
            return {}
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            return {}

    def poll_boundary_template_status(self, tenant_id: str, hierarchy_type: str, max_attempts: int = 30, delay: int = 2) -> Dict:
        """Poll for boundary template generation completion

        Args:
            tenant_id: Tenant ID
            hierarchy_type: Hierarchy type
            max_attempts: Maximum polling attempts
            delay: Delay between attempts (seconds)

        Returns:
            Dict with fileStoreId when complete
        """
        url = f"{self.boundary_mgmt_url}/v1/_generate-search"

        params = {
            "tenantId": tenant_id,
            "hierarchyType": hierarchy_type
        }

        # Override userInfo tenantId to match the request tenant
        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant_id

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": user_info_copy,
                "msgId": f"{int(time.time() * 1000)}|en_IN",
                "plainAccessRequest": {}
            }
        }


        headers = {'Content-Type': 'application/json'}

        print(f"\n⏳ Polling for template generation (max {max_attempts} attempts)...")

        for attempt in range(1, max_attempts + 1):
            try:
                response = requests.post(url, json=payload, headers=headers, params=params)
               
                response.raise_for_status()
                data = response.json()

                resources = data.get('GeneratedResource', [])
                if resources:
                    resource = resources[0]
                    status = resource.get('status')
                    filestore_id = resource.get('fileStoreid')
                    error_details = resource.get('error', 'No error details provided')
                    additional_details = resource.get('additionalDetails', {})

                    print(f"   Attempt {attempt}/{max_attempts}: Status = {status}")

                    if status == 'completed' and filestore_id:
                        print(f"\n✅ Template generation complete!")
                        print(f"   FileStore ID: {filestore_id}")
                        return resource
                    elif status == 'failed':
                        print(f"\n❌ Template generation failed")
                        print(f"   Error: {error_details}")
                        if additional_details:
                            print(f"   Additional Details: {json.dumps(additional_details, indent=2)[:500]}")
                        print(f"\n📋 Full resource response:")
                        print(f"   {json.dumps(resource, indent=2)[:1000]}")
                        return resource

                time.sleep(delay)

            except Exception as e:
                print(f"   Attempt {attempt}/{max_attempts}: Error - {str(e)[:100]}")
                time.sleep(delay)

        print(f"\n⚠️ Template generation timed out after {max_attempts} attempts")
        return {}

    def download_boundary_template(self, tenant_id: str, filestore_id: str, hierarchy_type: str = "ADMIN", output_path: str = None, return_url: bool = False):
        """Download boundary template from filestore

        Args:
            tenant_id: Tenant ID
            filestore_id: FileStore ID
            hierarchy_type: Hierarchy type for filename (optional)
            output_path: Path to save file (optional)
            return_url: If True, return dict with both path and download URL

        Returns:
            Path to downloaded file OR dict with 'path' and 'url' if return_url=True
        """
        import os
        url = f"{self.filestore_url}/v1/files/url"

        params = {
            "tenantId": tenant_id,
            "fileStoreIds": filestore_id
        }

        try:
            response = requests.get(url, params=params)
            response.raise_for_status()
            data = response.json()

            # Response format: { "fileStoreIds": [{"id": "xxx", "url": "s3://..."}] }
            file_urls = data.get('fileStoreIds', [])
            if not file_urls:
                print("❌ No file URL found in response")
                return None

            file_url = file_urls[0].get('url')
            if not file_url:
                print("❌ Invalid file URL")
                return None

            print(f"\n📥 Downloading from S3...")

            # Download the file
            file_response = requests.get(file_url)
            file_response.raise_for_status()

            # Determine output path
            if not output_path:
                os.makedirs('templates/boundary', exist_ok=True)
                output_path = f'templates/boundary/boundary_template_{tenant_id}_{hierarchy_type}.xlsx'

            with open(output_path, 'wb') as f:
                f.write(file_response.content)

            print(f"✅ Template downloaded: {output_path}")
            print(f"📊 File size: {len(file_response.content)} bytes")

            if return_url:
                return {
                    'path': output_path,
                    'url': file_url
                }
            return output_path

        except Exception as e:
            print(f"❌ Download error: {str(e)[:200]}")
            return None

    def upload_file_to_filestore(self, file_path: str, tenant_id: str, module: str = "HCM-ADMIN-CONSOLE") -> str:
        """Upload file to filestore

        Args:
            file_path: Path to file to upload
            tenant_id: Tenant ID
            module: Module name

        Returns:
            FileStore ID of uploaded file
        """
        import os
        url = f"{self.filestore_url}/v1/files"

        try:
            with open(file_path, 'rb') as f:
                files = {'file': (os.path.basename(file_path), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {
                    'tenantId': tenant_id,
                    'module': module
                }

                print(f"\n📤 Uploading file: {os.path.basename(file_path)}")
                response = requests.post(url, files=files, data=data)
                response.raise_for_status()

                result = response.json()
                files_data = result.get('files', [])

                if files_data:
                    filestore_id = files_data[0].get('fileStoreId')
                    print(f"✅ File uploaded successfully!")
                    print(f"   FileStore ID: {filestore_id}")
                    return filestore_id
                else:
                    print("❌ No filestore ID in response")
                    return None

        except Exception as e:
            print(f"❌ Upload error: {str(e)[:200]}")
            return None

    def process_boundary_data(self, tenant_id: str, filestore_id: str = None, hierarchy_type: str = "ADMIN", action: str = "create", excel_file: str = None) -> Dict:
        """Process boundary data - creates boundaries one-by-one via direct API

        This method reads the boundary Excel file and creates boundaries directly
        via the boundary-service API, avoiding the need for bulk upload services.

        Args:
            tenant_id: Tenant ID
            filestore_id: FileStore ID (optional, for compatibility)
            hierarchy_type: Hierarchy type (default: ADMIN)
            action: Action type (create/update)
            excel_file: Path to Excel file with boundary data

        Returns:
            Dict with processing results
        """
        import pandas as pd

        results = {
            'status': 'processing',
            'boundaries_created': 0,
            'relationships_created': 0,
            'errors': []
        }

        if not excel_file:
            print("❌ No Excel file provided")
            results['status'] = 'failed'
            results['errors'].append("No Excel file provided")
            return results

        try:
            # Read boundary data from Excel
            print(f"\n📖 Reading boundary data from: {excel_file}")

            # Try different sheet names
            sheet_name = 'Boundary'
            try:
                df = pd.read_excel(excel_file, sheet_name='Boundary')
            except:
                try:
                    df = pd.read_excel(excel_file, sheet_name='Boundary Data')
                except:
                    df = pd.read_excel(excel_file, sheet_name=0)  # First sheet

            print(f"   Found {len(df)} boundary records")
            print(f"   Columns: {list(df.columns)}")

            # Get hierarchy definition to understand boundary types
            hierarchy = self._get_boundary_hierarchy(tenant_id, hierarchy_type)
            if hierarchy:
                boundary_types = [h['boundaryType'] for h in hierarchy.get('boundaryHierarchy', [])]
                print(f"   Hierarchy: {' → '.join(boundary_types)}")
            else:
                print("   ⚠️ Could not fetch hierarchy, will use boundaryType from Excel")
                boundary_types = df['boundaryType'].unique().tolist() if 'boundaryType' in df.columns else []

            # Check if Excel has the standard format (code, name, boundaryType, parentCode)
            if 'code' in df.columns and 'boundaryType' in df.columns:
                print("   Using standard format (code, boundaryType, parentCode)")

                # Map Excel boundary types to hierarchy types if they don't match
                # Common mappings for Punjab-style templates
                type_mapping = {
                    'State': 'Country',  # If hierarchy starts with Country
                    'District': 'State',
                    'Tehsil': 'City',
                    'Block': 'Ward',
                    'Village': 'Locality'
                }

                # Check if we need mapping (Excel types vs hierarchy types)
                excel_types = set(df['boundaryType'].unique())
                hierarchy_set = set(boundary_types) if boundary_types else set()

                # If Excel types match hierarchy, no mapping needed
                if excel_types.issubset(hierarchy_set) or not hierarchy_set:
                    use_mapping = False
                    print("   Boundary types match hierarchy - no mapping needed")
                else:
                    use_mapping = True
                    print(f"   ⚠️ Boundary type mismatch detected")
                    print(f"      Excel types: {excel_types}")
                    print(f"      Hierarchy types: {hierarchy_set}")
                    print(f"      Will attempt to map types")

                # Process each row
                for idx, row in df.iterrows():
                    code = str(row.get('code', '')).strip()
                    boundary_type = str(row.get('boundaryType', '')).strip()
                    parent_code = str(row.get('parentCode', '')).strip() if pd.notna(row.get('parentCode')) else None

                    if not code or not boundary_type:
                        continue

                    # Apply type mapping if needed
                    mapped_type = type_mapping.get(boundary_type, boundary_type) if use_mapping else boundary_type

                    # Create boundary entity
                    success = self._create_boundary_entity(tenant_id, code)
                    if success:
                        results['boundaries_created'] += 1

                    # Create boundary relationship - try with original type first, then mapped
                    rel_success = self._create_boundary_relationship(
                        tenant_id, hierarchy_type, code, boundary_type, parent_code
                    )
                    if not rel_success and use_mapping and mapped_type != boundary_type:
                        # Try with mapped type
                        rel_success = self._create_boundary_relationship(
                            tenant_id, hierarchy_type, code, mapped_type, parent_code
                        )
                    if rel_success:
                        results['relationships_created'] += 1
            else:
                # Handle column-per-level format
                print("   Using column-per-level format")
                for boundary_type in boundary_types:
                    if boundary_type not in df.columns:
                        continue

                    boundaries_at_level = df[boundary_type].dropna().unique()

                    for boundary_code in boundaries_at_level:
                        if pd.isna(boundary_code) or str(boundary_code).strip() == '':
                            continue

                        boundary_code = str(boundary_code).strip()
                        success = self._create_boundary_entity(tenant_id, boundary_code)
                        if success:
                            results['boundaries_created'] += 1

                        parent_type_idx = boundary_types.index(boundary_type) - 1
                        parent_code = None
                        if parent_type_idx >= 0:
                            parent_type = boundary_types[parent_type_idx]
                            row = df[df[boundary_type] == boundary_code].iloc[0] if len(df[df[boundary_type] == boundary_code]) > 0 else None
                            if row is not None and parent_type in df.columns:
                                parent_code = str(row[parent_type]).strip() if pd.notna(row[parent_type]) else None

                        rel_success = self._create_boundary_relationship(
                            tenant_id, hierarchy_type, boundary_code, boundary_type, parent_code
                        )
                        if rel_success:
                            results['relationships_created'] += 1

            results['status'] = 'completed'
            print(f"\n✅ Boundary processing completed!")
            print(f"   Boundaries created: {results['boundaries_created']}")
            print(f"   Relationships created: {results['relationships_created']}")

        except Exception as e:
            print(f"❌ Error processing boundaries: {str(e)}")
            results['status'] = 'failed'
            results['errors'].append(str(e))

        return results

    def _get_boundary_hierarchy(self, tenant_id: str, hierarchy_type: str) -> Dict:
        """Fetch boundary hierarchy definition"""
        url = f"{self.boundary_url}/boundary-hierarchy-definition/_search"

        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant_id

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": user_info_copy
            },
            "BoundaryTypeHierarchySearchCriteria": {
                "tenantId": tenant_id,
                "hierarchyType": hierarchy_type
            }
        }

        try:
            response = requests.post(url, json=payload, headers={'Content-Type': 'application/json'})
            response.raise_for_status()
            data = response.json()
            hierarchies = data.get('BoundaryHierarchy', [])
            return hierarchies[0] if hierarchies else None
        except Exception as e:
            print(f"   ⚠️ Error fetching hierarchy: {str(e)[:100]}")
            return None

    def _create_boundary_entity(self, tenant_id: str, code: str) -> bool:
        """Create a single boundary entity"""
        url = f"{self.boundary_url}/boundary/_create"

        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant_id

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "msgId": f"create-boundary-{int(time.time()*1000)}",
                "authToken": self.auth_token,
                "userInfo": user_info_copy
            },
            "Boundary": [{
                "tenantId": tenant_id,
                "code": code,
                "geometry": {"type": "Polygon", "coordinates": [[[0,0],[0,1],[1,1],[1,0],[0,0]]]}
            }]
        }

        try:
            response = requests.post(url, json=payload, headers={'Content-Type': 'application/json'})
            if response.status_code in [200, 201, 202]:
                print(f"   ✅ Created boundary: {code}")
                return True
            else:
                data = response.json()
                error_code = data.get('Errors', [{}])[0].get('code', '')
                error_msg = data.get('Errors', [{}])[0].get('message', '')
                if error_code == 'DUPLICATE_CODE' or 'already exists' in str(error_msg).lower():
                    print(f"   ⚠️ Boundary exists: {code}")
                    return True  # Already exists is OK
                else:
                    print(f"   ❌ Failed to create boundary {code}: {error_code or error_msg or response.status_code}")
                    return False
        except Exception as e:
            print(f"   ❌ Error creating boundary {code}: {str(e)[:100]}")
            return False

    def _create_boundary_relationship(self, tenant_id: str, hierarchy_type: str,
                                       code: str, boundary_type: str, parent_code: str = None) -> bool:
        """Create boundary relationship (parent-child link)"""
        url = f"{self.boundary_url}/boundary-relationships/_create"

        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant_id

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "msgId": f"create-relationship-{int(time.time()*1000)}",
                "authToken": self.auth_token,
                "userInfo": user_info_copy
            },
            "BoundaryRelationship": {
                "tenantId": tenant_id,
                "hierarchyType": hierarchy_type,
                "code": code,
                "boundaryType": boundary_type
            }
        }

        if parent_code:
            payload["BoundaryRelationship"]["parent"] = parent_code

        try:
            response = requests.post(url, json=payload, headers={'Content-Type': 'application/json'})
            if response.status_code in [200, 201, 202]:
                parent_info = f" (parent: {parent_code})" if parent_code else " (root)"
                print(f"   ✅ Created relationship: {code} [{boundary_type}]{parent_info}")
                return True
            else:
                data = response.json()
                error_code = data.get('Errors', [{}])[0].get('code', '')
                error_msg = data.get('Errors', [{}])[0].get('message', '')
                if 'already exists' in str(error_msg).lower() or error_code == 'DUPLICATE':
                    print(f"   ⚠️ Relationship exists: {code}")
                    return True
                else:
                    print(f"   ❌ Failed relationship {code}: {error_msg[:80] if error_msg else error_code or response.status_code}")
                    return False
        except Exception as e:
            print(f"   ❌ Error creating relationship {code}: {str(e)[:100]}")
            return False

    def delete_all_boundaries(self, tenant_id: str) -> Dict:
        """Delete all boundary entities for a tenant

        Args:
            tenant_id: Tenant ID (e.g., 'statea', 'pg.citya')

        Returns:
            dict: {deleted: count, failed: count}
        """
        print(f"\n🗑️ Deleting all boundaries for tenant: {tenant_id}")

        results = {'deleted': 0, 'failed': 0, 'codes': []}

        # Search for all boundaries
        url = f"{self.boundary_url}/boundary/_search"
        payload = {
            "RequestInfo": {
                "apiId": "asset-services",
                "msgId": f"search-{int(time.time()*1000)}",
                "authToken": self.auth_token,
                "userInfo": self.user_info
            },
            "BoundaryCriteria": {
                "tenantId": tenant_id,
                "limit": 500,
                "offset": 0
            }
        }

        try:
            response = requests.post(url, json=payload, headers={'Content-Type': 'application/json'})
            if response.status_code != 200:
                print(f"   ❌ Failed to search boundaries: {response.status_code}")
                return results

            data = response.json()
            boundaries = data.get('Boundary', [])

            if not boundaries:
                print(f"   ℹ️ No boundaries found for tenant {tenant_id}")
                return results

            print(f"   Found {len(boundaries)} boundaries to delete")

            # Delete each boundary
            for boundary in boundaries:
                code = boundary.get('code')
                if not code:
                    continue

                delete_url = f"{self.boundary_url}/boundary/_delete"
                delete_payload = {
                    "RequestInfo": {
                        "apiId": "asset-services",
                        "msgId": f"delete-{int(time.time()*1000)}",
                        "authToken": self.auth_token,
                        "userInfo": self.user_info
                    }
                }

                try:
                    del_response = requests.post(
                        f"{delete_url}?tenantId={tenant_id}&code={code}",
                        json=delete_payload,
                        headers={'Content-Type': 'application/json'}
                    )

                    if del_response.status_code == 200:
                        print(f"   ✅ Deleted: {code}")
                        results['deleted'] += 1
                        results['codes'].append(code)
                    else:
                        print(f"   ❌ Failed to delete {code}: {del_response.status_code}")
                        results['failed'] += 1
                except Exception as e:
                    print(f"   ❌ Error deleting {code}: {str(e)[:30]}")
                    results['failed'] += 1

        except Exception as e:
            print(f"   ❌ Error searching boundaries: {str(e)[:50]}")

        print(f"\n   Summary: Deleted {results['deleted']}, Failed {results['failed']}")
        return results

    def delete_boundary_hierarchy(self, tenant_id: str, hierarchy_type: str) -> Dict:
        """Delete a boundary hierarchy definition

        Args:
            tenant_id: Tenant ID
            hierarchy_type: Hierarchy type (e.g., 'REVENUE', 'ADMIN')

        Returns:
            dict: {status: 'success'|'failed', message: str}
        """
        print(f"\n🗑️ Deleting hierarchy {hierarchy_type} for tenant: {tenant_id}")

        url = f"{self.boundary_url}/boundary-hierarchy-definition/_delete"

        payload = {
            "RequestInfo": {
                "apiId": "asset-services",
                "msgId": f"delete-hierarchy-{int(time.time()*1000)}",
                "authToken": self.auth_token,
                "userInfo": self.user_info
            },
            "BoundaryTypeHierarchyDefinition": {
                "tenantId": tenant_id,
                "hierarchyType": hierarchy_type
            }
        }

        try:
            response = requests.post(url, json=payload, headers={'Content-Type': 'application/json'})
            if response.status_code == 200:
                print(f"   ✅ Deleted hierarchy: {hierarchy_type}")
                return {'status': 'success', 'message': f'Deleted {hierarchy_type}'}
            else:
                error = response.json().get('Errors', [{}])[0].get('message', response.text[:100])
                print(f"   ❌ Failed: {error}")
                return {'status': 'failed', 'message': error}
        except Exception as e:
            print(f"   ❌ Error: {str(e)[:50]}")
            return {'status': 'failed', 'message': str(e)}

    # ========================================================================
    # HRMS EMPLOYEE METHODS
    # ========================================================================

    def fetch_departments(self, tenant: str) -> List[Dict]:
        """Fetch all departments from MDMS

        Args:
            tenant: Tenant ID

        Returns:
            List of department objects with code and name
        """
        print(f"📥 Fetching departments from MDMS for tenant: {tenant}")
        departments = self.search_mdms_data(schema_code='common-masters.Department', tenant=tenant)
        print(f"   ✅ Found {len(departments)} department(s)")
        return departments

    def fetch_designations(self, tenant: str) -> List[Dict]:
        """Fetch all designations from MDMS

        Args:
            tenant: Tenant ID

        Returns:
            List of designation objects with code and name
        """
        print(f"📥 Fetching designations from MDMS for tenant: {tenant}")
        designations = self.search_mdms_data(schema_code='common-masters.Designation', tenant=tenant)
        print(f"   ✅ Found {len(designations)} designation(s)")
        return designations

    def fetch_roles(self, tenant: str) -> List[Dict]:
        """Fetch all roles from MDMS

        Args:
            tenant: Tenant ID

        Returns:
            List of role objects with code and name
        """
        try:
            print(f"📥 Fetching roles from MDMS for tenant: {tenant}")

            # Try to fetch from MDMS roles schema
            roles = self.search_mdms_data(
                schema_code='ACCESSCONTROL-ROLES.roles',
                tenant=tenant,
                limit=200
            )

            if roles and len(roles) > 0:
                print(f"   ✅ Found {len(roles)} role(s) from MDMS")
                return roles
            else:
                print(f"   ⚠️  No roles found in MDMS, using defaults")
                # Return default PGR roles
                return self._get_default_roles()

        except Exception as e:
            print(f"   ⚠️  Error fetching roles from MDMS: {str(e)[:150]}")
            print(f"   ℹ️  Using default PGR roles")
            return self._get_default_roles()

    def _get_default_roles(self) -> List[Dict]:
        """Get default roles based on default-data-handler HRMS.json

        Returns:
            List of default role objects commonly used in eGov systems
        """
        return [
            # Report Viewer
            {"code": "TICKET_REPORT_VIEWER", "name": "Report Viewer", "description": "One who will view the reports of tickets"},

            # PGR Roles (from default-data-handler ACCESSCONTROL-ROLES.roles.json)
            {"code": "PGR_LME", "name": "Complaint Resolver", "description": "One who will resolve complaints"},
            {"code": "GRO", "name": "Complaint Assessor", "description": "One who will assess & assign complaints"},
            {"code": "CSR", "name": "Complainant", "description": "One who will create complaints"},
            {"code": "PGR_VIEWER", "name": "PGR Viewer role", "description": " "},

            # Admin Roles
            {"code": "LOC_ADMIN", "name": "Localisation admin", "description": "LOC_ADMIN"},
            {"code": "MDMS_ADMIN", "name": "MDMS ADMIN", "description": "MDMS User that can create and search schema"},
            {"code": "HRMS_ADMIN", "name": "HRMS Admin", "description": "HRMS Admin"},
            {"code": "WORKFLOW_ADMIN", "name": "WORKFLOW ADMIN", "description": "WORKFLOW User that can create and search Workflow"},
            {"code": "SUPERUSER", "name": "Super User", "description": "System Administrator. Can change all master data and has access to all the system screens."},

            # Common Roles
            {"code": "EMPLOYEE", "name": "Employee", "description": "Default role for all employees"},
            {"code": "CITIZEN", "name": "Citizen", "description": "Citizen who can raise complaint"},
            {"code": "ANONYMOUS", "name": "Anonymous User", "description": "Anonymous User to be used in case of no auth"},
            {"code": "COMMON_EMPLOYEE", "name": "Basic employee roles", "description": "Basic employee roles"},

            # System Roles
            {"code": "REINDEXING_ROLE", "name": "Reindexing Role", "description": "Role for reindexing for encrypted data access"},
            {"code": "INTERNAL_MICROSERVICE_ROLE", "name": "Internal Microservice Role", "description": "Internal role for plain access"},
            {"code": "SYSTEM", "name": "System user", "description": "System user role"},
            {"code": "QA_AUTOMATION", "name": "QA Automation", "description": "QA Automation"},

            # Escalation Roles
            {"code": "SUPERVISOR", "name": "Auto Escalation Supervisor", "description": "Escalation to particular role"},
            {"code": "AUTO_ESCALATE", "name": "Auto Escalation Employee", "description": "Auto Escalation Employee"}
        ]

    def ensure_roles_in_mdms(self, tenant: str, auto_create: bool = True) -> bool:
        """Ensure all required roles exist in MDMS, optionally creating them if missing

        Args:
            tenant: Tenant ID
            auto_create: If True, automatically create missing roles in MDMS

        Returns:
            True if all roles exist or were created successfully, False otherwise
        """
        print(f"\n🔍 Checking roles in MDMS for tenant: {tenant}")

        # Fetch existing roles from MDMS
        try:
            existing_roles = self.search_mdms_data(
                schema_code='ACCESSCONTROL-ROLES.roles',
                tenant=tenant,
                limit=200
            )
            existing_codes = {role.get('code') for role in existing_roles if role.get('code')}
            print(f"   ✅ Found {len(existing_codes)} existing roles in MDMS")
        except Exception as e:
            print(f"   ⚠️  Could not fetch existing roles: {str(e)[:150]}")
            existing_codes = set()

        # Get default roles that should exist
        default_roles = self._get_default_roles()
        required_codes = {role.get('code') for role in default_roles if role.get('code')}

        # Find missing roles
        missing_codes = required_codes - existing_codes

        if not missing_codes:
            print(f"   ✅ All {len(required_codes)} required roles already exist in MDMS")
            return True

        print(f"   ⚠️  Missing {len(missing_codes)} roles: {', '.join(sorted(missing_codes))}")

        if not auto_create:
            print(f"   ℹ️  Auto-create is disabled. Roles must be created manually.")
            return False

        # Create missing roles
        print(f"\n📤 Creating {len(missing_codes)} missing roles in MDMS...")
        created_count = 0
        failed_roles = []

        for role in default_roles:
            role_code = role.get('code')
            if role_code not in missing_codes:
                continue

            try:
                # Prepare MDMS data structure
                mdms_data = {
                    "code": role.get('code'),
                    "name": role.get('name'),
                    "description": role.get('description', '')
                }

                # Create in MDMS
                self.create_mdms_data(
                    schema_code='ACCESSCONTROL-ROLES.roles',
                    data=mdms_data,
                    tenant=tenant,
                    unique_identifier=role_code
                )
                created_count += 1
                print(f"   ✅ Created role: {role_code} ({role.get('name')})")

            except Exception as e:
                error_msg = str(e)
                # Check if it's a duplicate error (role was created between our check and now)
                if 'DUPLICATE' in error_msg.upper() or 'already exists' in error_msg.lower():
                    print(f"   ℹ️  Role already exists: {role_code}")
                    created_count += 1
                else:
                    print(f"   ❌ Failed to create role {role_code}: {error_msg[:150]}")
                    failed_roles.append(role_code)

        if failed_roles:
            print(f"\n⚠️  Failed to create {len(failed_roles)} roles: {', '.join(failed_roles)}")
            return False

        print(f"\n✅ Successfully ensured all {len(required_codes)} roles exist in MDMS ({created_count} created)")
        return True

    def fetch_boundaries(self, tenant: str, hierarchy_type: str = "ADMIN") -> List[Dict]:
        """Fetch boundaries from boundary service v2

        Args:
            tenant: Tenant ID
            hierarchy_type: Hierarchy type (default: ADMIN)

        Returns:
            List of boundary objects with code and type
        """
        url = f"{self.boundary_url}/boundary/_search"

        # Override userInfo tenantId
        user_info_copy = self.user_info.copy()
        user_info_copy['tenantId'] = tenant

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": user_info_copy,
                "msgId": f"{int(time.time() * 1000)}|en_IN"
            },
            "Boundary": {
                "tenantId": tenant,
                "hierarchyType": hierarchy_type,
                "limit": 100,
                "offset": 0
            }
        }

        headers = {'Content-Type': 'application/json'}

        try:
            print(f"📥 Fetching boundaries from boundary service for tenant: {tenant}")
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            data = response.json()

            # Extract boundaries from response
            # API returns: {"TenantBoundary": [{"hierarchyType": "...", "boundary": [...]}]}
            tenant_boundaries = data.get('TenantBoundary', [])

            boundaries = []
            for tb in tenant_boundaries:
                boundary_list = tb.get('boundary', {}).get('children', [])
                for boundary in boundary_list:
                    boundaries.append({
                        'code': boundary.get('code', ''),
                        'name': boundary.get('name', ''),
                        'boundaryType': boundary.get('boundaryType', 'City'),
                        'label': boundary.get('label', ''),
                        'latitude': boundary.get('latitude'),
                        'longitude': boundary.get('longitude')
                    })

            print(f"   ✅ Found {len(boundaries)} boundarie(s)")

            # If no boundaries found, return default
            if not boundaries:
                boundaries = [{"code": tenant.split('.')[0], "name": tenant.split('.')[0], "boundaryType": "City"}]

            return boundaries

        except Exception as e:
            print(f"   ⚠️  Could not fetch boundaries: {str(e)[:200]}")
            # Return default boundary
            return [{"code": tenant.split('.')[0], "name": tenant.split('.')[0], "boundaryType": "City"}]

    def fetch_gender_types(self, tenant: str) -> list:
        """Fetch Gender types from MDMS

        Args:
            tenant: Tenant ID

        Returns:
            List of gender codes
        """
        try:
            print(f"📥 Fetching Gender types from MDMS...")
            gender_data = self.search_mdms_data(
                schema_code='common-masters.GenderType',
                tenant=tenant
            )

            genders = [g.get('code', '') for g in gender_data if g.get('active', True)]
            print(f"   ✅ Found {len(genders)} gender type(s)")
            return genders if genders else ['MALE', 'FEMALE', 'TRANSGENDER', 'OTHERS']

        except Exception as e:
            print(f"   ⚠️  Using default gender types: {str(e)[:100]}")
            return ['MALE', 'FEMALE', 'TRANSGENDER', 'OTHERS']

    def fetch_employee_status(self, tenant: str) -> list:
        """Fetch Employee Status from MDMS

        Args:
            tenant: Tenant ID

        Returns:
            List of employee status codes
        """
        try:
            print(f"📥 Fetching Employee Status from MDMS...")
            status_data = self.search_mdms_data(
                schema_code='egov-hrms.EmployeeStatus',
                tenant=tenant
            )

            statuses = [s.get('code', '') for s in status_data if s.get('active', True)]
            print(f"   ✅ Found {len(statuses)} employee status(es)")
            return statuses if statuses else ['EMPLOYED', 'SUSPENDED', 'RETIRED', 'TERMINATED']

        except Exception as e:
            print(f"   ⚠️  Using default employee statuses: {str(e)[:100]}")
            return ['EMPLOYED', 'SUSPENDED', 'RETIRED', 'TERMINATED']

    def fetch_employee_types(self, tenant: str) -> list:
        """Fetch Employee Types from MDMS

        Args:
            tenant: Tenant ID

        Returns:
            List of employee type codes
        """
        try:
            print(f"📥 Fetching Employee Types from MDMS...")
            type_data = self.search_mdms_data(
                schema_code='egov-hrms.EmployeeType',
                tenant=tenant
            )

            types = [t.get('code', '') for t in type_data if t.get('active', True)]
            print(f"   ✅ Found {len(types)} employee type(s)")
            return types if types else ['PERMANENT', 'TEMPORARY', 'CONTRACT', 'DAILYWAGES']

        except Exception as e:
            print(f"   ⚠️  Using default employee types: {str(e)[:100]}")
            return ['PERMANENT', 'TEMPORARY', 'CONTRACT', 'DAILYWAGES']

    def fetch_hierarchy_types(self, tenant: str) -> list:
        """Fetch Hierarchy Types from Boundary Service

        Args:
            tenant: Tenant ID

        Returns:
            List of unique hierarchy type codes
        """
        try:
            print(f"📥 Fetching Hierarchy Types from Boundary Service...")
            url = f"{self.boundary_url}/boundary-hierarchy-definition/_search"


            # Override userInfo tenantId
            user_info_copy = self.user_info.copy()
            user_info_copy['tenantId'] = tenant

            payload = {
                "RequestInfo": {
                    "apiId": "Rainmaker",
                    "authToken": self.auth_token,
                    "userInfo": user_info_copy,
                },
                "BoundaryTypeHierarchySearchCriteria": {
                    "tenantId": tenant
                }
            }


            headers = {'Content-Type': 'application/json'}
            response = requests.post(url, json=payload, headers=headers, timeout=30)
           
            response.raise_for_status()
            data = response.json()

            # Extract unique hierarchy types
            hierarchy_definitions = data.get('BoundaryHierarchy', [])
            hierarchy_types = list(set([h.get('hierarchyType', '') for h in hierarchy_definitions if h.get('hierarchyType')]))

            print(f"   ✅ Found {len(hierarchy_types)} hierarchy type(s)")
            return hierarchy_types if hierarchy_types else ['ADMIN', 'ADMIN1', 'ADMIN2', 'REVENUE']

        except Exception as e:
            print(f"   ⚠️  Using default hierarchy types: {str(e)[:100]}")
            return ['ADMIN', 'ADMIN1', 'ADMIN2', 'REVENUE']

    def generate_employee_template(self, tenant: str, output_path: str = None) -> str:
        """Generate dynamic Employee Master Excel template with dropdowns

        Fetches all data from MDMS and creates pre-filled template with:
        - Department dropdown (from MDMS)
        - Designation dropdown (from MDMS)
        - Role codes dropdown (from MDMS)
        - Employee Status dropdown (EMPLOYED, SUSPENDED, etc.)
        - Employee Type dropdown (PERMANENT, TEMPORARY, etc.)
        - Gender dropdown (Male, Female, Other)
        - Hierarchy Type dropdown (ADMIN, ADMIN1, etc.)
        - Boundary Type dropdown (from boundaries)
        - Boundary Code dropdown (from boundaries)

        Args:
            tenant: Tenant ID
            output_path: Output file path (default: templates/Employee_Master_Dynamic.xlsx)

        Returns:
            Path to generated template
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Protection
        from openpyxl.worksheet.datavalidation import DataValidation

        print("\n" + "="*70)
        print("   📋 GENERATING DYNAMIC EMPLOYEE TEMPLATE")
        print("="*70)

        # Fetch all data from MDMS and Boundary Service
        departments = self.fetch_departments(tenant)
        designations = self.fetch_designations(tenant)
        roles = self.fetch_roles(tenant)
        boundaries = self.fetch_boundaries(tenant)
        genders = self.fetch_gender_types(tenant)
        employee_statuses = self.fetch_employee_status(tenant)
        employee_types = self.fetch_employee_types(tenant)
        hierarchy_types = self.fetch_hierarchy_types(tenant)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Master"

        # Create reference sheets for dropdowns
        ref_dept = wb.create_sheet("Ref_Departments")
        ref_desig = wb.create_sheet("Ref_Designations")
        ref_roles = wb.create_sheet("Ref_Roles")
        ref_boundaries = wb.create_sheet("Ref_Boundaries")

        # Hide reference sheets
        ref_dept.sheet_state = 'hidden'
        ref_desig.sheet_state = 'hidden'
        ref_roles.sheet_state = 'hidden'
        ref_boundaries.sheet_state = 'hidden'

        # Populate reference sheets
        # Departments
        ref_dept['A1'] = 'Department Code'
        ref_dept['B1'] = 'Department Name'
        for i, dept in enumerate(departments, 2):
            ref_dept[f'A{i}'] = dept.get('code', '')
            ref_dept[f'B{i}'] = dept.get('name', '')

        # Designations
        ref_desig['A1'] = 'Designation Code'
        ref_desig['B1'] = 'Designation Name'
        for i, desig in enumerate(designations, 2):
            ref_desig[f'A{i}'] = desig.get('code', '')
            ref_desig[f'B{i}'] = desig.get('name', '')

        # Roles
        ref_roles['A1'] = 'Role Code'
        ref_roles['B1'] = 'Role Name'
        for i, role in enumerate(roles, 2):
            ref_roles[f'A{i}'] = role.get('code', '')
            ref_roles[f'B{i}'] = role.get('name', '')

        # Boundaries
        ref_boundaries['A1'] = 'Boundary Code'
        ref_boundaries['B1'] = 'Boundary Type'
        for i, boundary in enumerate(boundaries, 2):
            ref_boundaries[f'A{i}'] = boundary.get('code', '')
            ref_boundaries[f'B{i}'] = boundary.get('boundaryType', 'City')

        # Define headers (user only sees names, codes are handled internally)
        headers = [
            'User Name*',
            'Mobile Number*',
            'Password',
            'Department Name*',
            'Designation Name*',
            'Role Names (comma separated)*',
            'Employee Status',
            'Employee Type',
            'Gender',
            'Hierarchy Type',
            'Boundary Type',
            'Boundary Code',
            'Assignment From Date*',
            'Date of Appointment*'
        ]

        # Write headers
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Add comment to Role Names column
            if header == 'Role Names (comma separated)*':
                from openpyxl.comments import Comment
                # Get list of available role names for the tooltip
                available_roles = ', '.join([r.get('name', '') for r in roles[:10]])
                if len(roles) > 10:
                    available_roles += f', ... ({len(roles)} total roles)'

                comment = Comment(
                    f"💡 Enter comma-separated role names\n\n"
                    f"Example: Employee,PGR Viewer,PGR Last Mile Employee\n\n"
                    f"Available roles:\n{available_roles}\n\n"
                    f"See 'Ref_Roles' sheet for complete list",
                    "System"
                )
                cell.comment = comment

        # Create data validations with dropdowns using NAMES (not codes)
        # Department Name dropdown (Column D)
        dept_names = ','.join([d.get('name', '') for d in departments[:50]])  # Excel limit
        if len(departments) <= 50:
            dv_dept = DataValidation(type="list", formula1=f'"{dept_names}"', allow_blank=True)
        else:
            dv_dept = DataValidation(type="list", formula1=f"Ref_Departments!$B$2:$B${len(departments)+1}", allow_blank=True)
        dv_dept.error = 'Please select a valid department name'
        dv_dept.errorTitle = 'Invalid Department'
        ws.add_data_validation(dv_dept)
        dv_dept.add(f'D2:D1000')

        # Designation Name dropdown (Column E)
        desig_names = ','.join([d.get('name', '') for d in designations[:50]])
        if len(designations) <= 50:
            dv_desig = DataValidation(type="list", formula1=f'"{desig_names}"', allow_blank=True)
        else:
            dv_desig = DataValidation(type="list", formula1=f"Ref_Designations!$B$2:$B${len(designations)+1}", allow_blank=True)
        dv_desig.error = 'Please select a valid designation name'
        dv_desig.errorTitle = 'Invalid Designation'
        ws.add_data_validation(dv_desig)
        dv_desig.add(f'E2:E1000')

        # Employee Status dropdown (Column G) - Dynamic from MDMS
        status_list = ','.join(employee_statuses)
        dv_status = DataValidation(type="list", formula1=f'"{status_list}"', allow_blank=True)
        dv_status.error = f'Please select from: {status_list}'
        dv_status.errorTitle = 'Invalid Employee Status'
        ws.add_data_validation(dv_status)
        dv_status.add(f'G2:G1000')

        # Employee Type dropdown (Column H) - Dynamic from MDMS
        type_list = ','.join(employee_types)
        dv_type = DataValidation(type="list", formula1=f'"{type_list}"', allow_blank=True)
        dv_type.error = f'Please select from: {type_list}'
        dv_type.errorTitle = 'Invalid Employee Type'
        ws.add_data_validation(dv_type)
        dv_type.add(f'H2:H1000')

        # Gender dropdown (Column I) - Dynamic from MDMS
        gender_list = ','.join(genders)
        dv_gender = DataValidation(type="list", formula1=f'"{gender_list}"', allow_blank=True)
        dv_gender.error = f'Please select from: {gender_list}'
        dv_gender.errorTitle = 'Invalid Gender'
        ws.add_data_validation(dv_gender)
        dv_gender.add(f'I2:I1000')

        # Hierarchy Type dropdown (Column J) - Dynamic from Boundary Service
        hierarchy_list = ','.join(hierarchy_types)
        dv_hierarchy = DataValidation(type="list", formula1=f'"{hierarchy_list}"', allow_blank=True)
        dv_hierarchy.error = f'Please select from: {hierarchy_list}'
        dv_hierarchy.errorTitle = 'Invalid Hierarchy Type'
        ws.add_data_validation(dv_hierarchy)
        dv_hierarchy.add(f'J2:J1000')

        # Boundary Type dropdown (Column K)
        boundary_types = ','.join(list(set([b.get('boundaryType', 'City') for b in boundaries])))
        dv_boundary_type = DataValidation(type="list", formula1=f'"{boundary_types}"', allow_blank=True)
        ws.add_data_validation(dv_boundary_type)
        dv_boundary_type.add(f'K2:K1000')

        # Boundary Code dropdown (Column L)
        boundary_codes = ','.join([b.get('code', '') for b in boundaries[:50]])
        if len(boundaries) <= 50:
            dv_boundary = DataValidation(type="list", formula1=f'"{boundary_codes}"', allow_blank=True)
        else:
            dv_boundary = DataValidation(type="list", formula1=f"Ref_Boundaries!$A$2:$A${len(boundaries)+1}", allow_blank=True)
        ws.add_data_validation(dv_boundary)
        dv_boundary.add(f'L2:L1000')

        # Set column widths
        column_widths = {
            'A': 20,  # User Name
            'B': 15,  # Mobile Number
            'C': 15,  # Password
            'D': 25,  # Department Name
            'E': 25,  # Designation Name
            'F': 50,  # Role Names (comma-separated)
            'G': 15,  # Employee Status
            'H': 15,  # Employee Type
            'I': 12,  # Gender
            'J': 15,  # Hierarchy Type
            'K': 15,  # Boundary Type
            'L': 15,  # Boundary Code
            'M': 25,  # Assignment From Date
            'N': 25,  # Date of Appointment
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Import datetime for date formatting
        from datetime import datetime

        # Add sample row with default values (using dynamic data from MDMS)
        ws['A2'] = 'Sample Employee'
        ws['B2'] = '9999999999'
        ws['C2'] = 'eGov@123'
        ws['D2'] = departments[0].get('name', 'Department 1') if departments else 'Department 1'
        ws['E2'] = designations[0].get('name', 'Designation 01') if designations else 'Designation 01'
        ws['F2'] = 'Employee,PGR Viewer'  # Role NAMES, not codes
        ws['G2'] = employee_statuses[0] if employee_statuses else 'EMPLOYED'
        ws['H2'] = employee_types[0] if employee_types else 'PERMANENT'
        ws['I2'] = genders[0] if genders else 'MALE'
        ws['J2'] = hierarchy_types[0] if hierarchy_types else 'ADMIN'
        ws['K2'] = 'City'
        ws['L2'] = tenant.split('.')[0]

        # Use actual Excel dates (not timestamps)
        ws['M2'] = datetime(2024, 9, 5)  # Assignment From Date
        ws['N2'] = datetime(2024, 6, 20)  # Date of Appointment

        # Format date columns
        from openpyxl.styles import numbers
        for row in range(2, 1001):  # Format rows 2-1000
            ws[f'M{row}'].number_format = numbers.FORMAT_DATE_XLSX14  # mm-dd-yy
            ws[f'N{row}'].number_format = numbers.FORMAT_DATE_XLSX14

        # Add instructions sheet
        instructions = wb.create_sheet("Instructions", 0)
        instructions['A1'] = "📋 EMPLOYEE MASTER TEMPLATE - INSTRUCTIONS"
        instructions['A1'].font = Font(bold=True, size=14, color='0066cc')

        instructions['A3'] = "✨ USER-FRIENDLY: Use NAMES everywhere! Codes are converted automatically."
        instructions['A5'] = "📝 Required Fields (marked with *):"
        instructions['A6'] = "  1. User Name* - Full name of the employee"
        instructions['A7'] = "  2. Mobile Number* - 10-digit mobile number"
        instructions['A8'] = "  3. Department Name* - Select from dropdown (e.g., 'Public Works')"
        instructions['A9'] = "  4. Designation Name* - Select from dropdown (e.g., 'Assistant Engineer')"
        instructions['A10'] = "  5. Role Names* - Type comma-separated role names (see 'README - Roles' sheet)"
        instructions['A11'] = "  6. Assignment From Date* - Use Excel date picker (e.g., 09/05/2024)"
        instructions['A12'] = "  7. Date of Appointment* - Use Excel date picker (e.g., 06/20/2024)"

        instructions['A14'] = "🎯 HOW TO SELECT MULTIPLE ROLES:"
        instructions['A15'] = "  1. Go to 'README - Roles' sheet"
        instructions['A16'] = "  2. Find the roles you want from the list"
        instructions['A17'] = "  3. Copy role names (NOT codes) - example: 'Employee,PGR Viewer,PGR Admin'"
        instructions['A18'] = "  4. Paste into the 'Role Names' column in Employee Master sheet"
        instructions['A19'] = "  "
        instructions['A20'] = "  💡 TIP: You can copy multiple role names at once and join with commas"
        instructions['A21'] = "  Example: Employee,PGR Viewer,PGR Last Mile Employee"

        instructions['A23'] = "💡 Optional Fields (with dropdowns):"
        instructions['A24'] = "  • Password - User password (default: eGov@123 if left blank)"
        instructions['A25'] = "  • Employee Status - EMPLOYED, SUSPENDED, RETIRED, TERMINATED (default: EMPLOYED)"
        instructions['A26'] = "  • Employee Type - PERMANENT, TEMPORARY, CONTRACT, DAILY_WAGE (default: PERMANENT)"
        instructions['A27'] = "  • Gender - Male, Female, Other, Transgender"
        instructions['A28'] = "  • Hierarchy Type - ADMIN, ADMIN1, ADMIN2, REVENUE (default: ADMIN)"
        instructions['A29'] = "  • Boundary Type - Select from available boundaries (default: City)"
        instructions['A30'] = "  • Boundary Code - Select from available boundaries (default: tenant code)"

        instructions['A32'] = "⚠️  IMPORTANT NOTES:"
        instructions['A33'] = "  1. USE NAMES NOT CODES - Department, Designation, and Role NAMES are used"
        instructions['A34'] = "  2. Auto-Conversion - System converts names to codes when uploading"
        instructions['A35'] = "  3. Employee Code - Auto-generated from User Name"
        instructions['A36'] = "  4. Jurisdiction - Automatically created using Boundary + Roles"
        instructions['A37'] = "  5. Roles - Same roles assigned to BOTH user and jurisdiction"
        instructions['A38'] = "  6. All dropdowns pre-filled from MDMS"
        instructions['A39'] = "  7. See 'README - Roles' sheet for complete list of available roles"
        instructions['A40'] = "  8. Delete sample row before uploading real data"
        instructions['A41'] = f"  9. Configured for tenant: {tenant}"

        instructions.column_dimensions['A'].width = 80

        # Create README - Roles sheet with all available roles
        readme_roles = wb.create_sheet("README - Roles")
        readme_roles['A1'] = "📚 AVAILABLE ROLES - Copy & Paste Role Names"
        readme_roles['A1'].font = Font(bold=True, size=14, color='0066cc')
        readme_roles['A2'] = f"Total Roles Available: {len(roles)}"
        readme_roles['A2'].font = Font(bold=True, color='006600')

        readme_roles['A4'] = "Instructions:"
        readme_roles['A4'].font = Font(bold=True, size=12, color='cc6600')
        readme_roles['A5'] = "  1. Find the roles you want to assign to the employee"
        readme_roles['A6'] = "  2. Copy the ROLE NAME (Column B) - NOT the code!"
        readme_roles['A7'] = "  3. For multiple roles, copy each name and join with commas"
        readme_roles['A8'] = "  4. Paste into 'Role Names' column in Employee Master sheet"
        readme_roles['A9'] = ""
        readme_roles['A10'] = "Example: Employee,PGR Viewer,PGR Last Mile Employee"
        readme_roles['A10'].font = Font(italic=True, color='0066cc')

        # Headers for role list
        readme_roles['A12'] = "Role Code"
        readme_roles['B12'] = "Role Name (COPY THIS)"
        readme_roles['C12'] = "Description"
        readme_roles['A12'].font = Font(bold=True, color='FFFFFF')
        readme_roles['B12'].font = Font(bold=True, color='FFFFFF')
        readme_roles['C12'].font = Font(bold=True, color='FFFFFF')

        from openpyxl.styles import PatternFill
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        readme_roles['A12'].fill = header_fill
        readme_roles['B12'].fill = header_fill
        readme_roles['C12'].fill = header_fill

        # Populate all roles
        row = 13
        for role in sorted(roles, key=lambda x: x.get('name', '')):
            readme_roles[f'A{row}'] = role.get('code', '')
            readme_roles[f'B{row}'] = role.get('name', '')
            readme_roles[f'C{row}'] = role.get('description', '')

            # Highlight role name column for easy copying
            readme_roles[f'B{row}'].font = Font(bold=True, color='006600')
            row += 1

        # Set column widths
        readme_roles.column_dimensions['A'].width = 35
        readme_roles.column_dimensions['B'].width = 45
        readme_roles.column_dimensions['C'].width = 60

        # Add note at bottom
        readme_roles[f'A{row+2}'] = "💡 Common Role Combinations:"
        readme_roles[f'A{row+2}'].font = Font(bold=True, size=11, color='cc6600')
        readme_roles[f'A{row+3}'] = "  • Basic Employee: Employee"
        readme_roles[f'A{row+4}'] = "  • PGR Complaint Resolver: PGR Last Mile Employee,PGR Viewer,Employee"
        readme_roles[f'A{row+5}'] = "  • PGR Complaint Assessor: Grievance Routing Officer,PGR Viewer,Employee"
        readme_roles[f'A{row+6}'] = "  • Admin User: HRMS Admin,MDMS ADMIN,Localisation admin,WORKFLOW ADMIN"

        # Save file
        if not output_path:
            os.makedirs('templates', exist_ok=True)
            output_path = f'templates/Employee_Master_Dynamic_{tenant}.xlsx'

        # Delete old file if exists to ensure clean generation
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
                print(f"\n🗑️  Deleted old template: {output_path}")
            except Exception as e:
                print(f"\n⚠️  Could not delete old file: {e}")

        wb.save(output_path)

        print("\n" + "="*70)
        print("  ✅ DYNAMIC TEMPLATE GENERATED SUCCESSFULLY")
        print("="*70)
        print(f"\n📄 File: {output_path}")
        print(f"\n📊 Statistics:")
        print(f"   • Departments: {len(departments)}")
        print(f"   • Designations: {len(designations)}")
        print(f"   • Roles: {len(roles)}")
        print(f"   • Boundaries: {len(boundaries)}")
        print(f"   • Dropdowns: 8 (Department, Designation, Status, Type, Gender, Hierarchy, Boundary Type, Boundary Code)")
        print("\n💡 Template includes:")
        print("   ✓ Employee Master sheet - Main data entry sheet")
        print("   ✓ Instructions sheet - Complete usage guide")
        print("   ✓ README - Roles sheet - All available roles with copy-paste ready names")
        print("   ✓ Hidden reference sheets (Departments, Designations, Roles, Boundaries)")
        print("   ✓ All dropdowns pre-filled from MDMS")
        print("   ✓ Sample row with default values")
        print("   ✓ Data validation for all dropdown fields")
        print("="*70)

        return output_path

    def create_employees(self, employee_list: List[Dict], tenant: str,
                        sheet_name: str = None, excel_file: str = None):
        """Bulk create employees via HRMS API with password update support

        Args:
            employee_list: List of employee objects
            tenant: Tenant ID
            sheet_name: Excel sheet name to update with status
            excel_file: Path to the uploaded Excel file

        Returns:
            Dict with creation results
        """
        # STEP 1: Ensure all required roles exist in MDMS before creating employees
        print(f"\n{'='*60}")
        print(f"🔐 PRE-CHECK: Validating Roles in MDMS")
        print(f"{'='*60}")

        roles_ok = self.ensure_roles_in_mdms(tenant=tenant, auto_create=True)

        if not roles_ok:
            error_msg = "⚠️  Cannot proceed: Some required roles are missing from MDMS and could not be created."
            print(f"\n{error_msg}")
            print(f"   Please ensure roles are created in MDMS before creating employees.")
            return {
                'created': 0,
                'exists': 0,
                'failed': len(employee_list),
                'errors': [error_msg]
            }

        # STEP 2: Proceed with employee creation
        create_url = f"{self.hrms_url}/employees/_create"
        update_url = f"{self.hrms_url}/employees/_update"
        search_url = f"{self.hrms_url}/employees/_search"

        results = {
            'created': 0,
            'exists': 0,
            'failed': 0,
            'errors': [],
            'password_updated': 0
        }

        print(f"\n{'='*60}")
        print(f"[UPLOADING] HRMS Employees")
        print(f"   Tenant: {tenant}")
        print(f"   Records: {len(employee_list)}")
        print(f"   API URL: {create_url}")
        print("="*60)

        # Track row-by-row status for Excel update
        row_statuses = []

        for i, employee in enumerate(employee_list, 1):
            emp_code = employee.get('code', str(i))

            # Extract custom password before creation (if provided)
            custom_password = employee.get('user', {}).get('password')
            needs_password_update = custom_password and custom_password not in ['eGov@123', None, '']

            # Override userInfo tenantId to match the request tenant
            user_info_copy = self.user_info.copy()
            user_info_copy['tenantId'] = tenant

            payload = {
                "RequestInfo": {
                    "apiId": "Rainmaker",
                    "ver": "1.0",
                    "action": "_create",
                    "msgId": f"{int(time.time() * 1000)}",
                    "authToken": self.auth_token,
                    "userInfo": user_info_copy
                },
                "Employees": [employee]
            }

            headers = {'Content-Type': 'application/json'}
            status = "SUCCESS"
            status_code = 200
            error_message = ""

            try:
                # STEP 2A: Create employee (system generates random password)
                response = requests.post(create_url, json=payload, headers=headers)
                status_code = response.status_code
                response.raise_for_status()

                created_data = response.json()
                created_employee = created_data.get('Employees', [{}])[0]

                print(f"   [OK] [{i}/{len(employee_list)}] {emp_code} - Created")
                results['created'] += 1

                # STEP 2B: Update password if custom password was provided
                if needs_password_update and created_employee:
                    try:
                        # Search for the created employee to get full details
                        search_payload = {
                            "RequestInfo": {
                                "apiId": "Rainmaker",
                                "authToken": self.auth_token,
                                "userInfo": user_info_copy,
                                "msgId": f"{int(time.time() * 1000)}|en_IN"
                            },
                            "codes": [emp_code],
                            "tenantId": tenant
                        }

                        # Add query parameters for search
                        search_params = {
                            "tenantId": tenant,
                            "codes": emp_code
                        }

                        search_response = requests.post(search_url, json=search_payload, headers=headers, params=search_params)
                        search_response.raise_for_status()
                        search_data = search_response.json()

                        full_employee = search_data.get('Employees', [{}])[0]

                        if full_employee and full_employee.get('id'):
                            # Update the password in the user object
                            full_employee['user']['password'] = custom_password

                            # Prepare update payload
                            update_payload = {
                                "RequestInfo": {
                                    "apiId": "Rainmaker",
                                    "ver": "1.0",
                                    "action": "_update",
                                    "msgId": f"{int(time.time() * 1000)}",
                                    "authToken": self.auth_token,
                                    "userInfo": user_info_copy
                                },
                                "Employees": [full_employee]
                            }

                            # Update employee with custom password
                            update_response = requests.post(update_url, json=update_payload, headers=headers)
                            update_response.raise_for_status()

                            print(f"   [✓] [{i}/{len(employee_list)}] {emp_code} - Password updated")
                            results['password_updated'] += 1
                        else:
                            print(f"   [⚠] [{i}/{len(employee_list)}] {emp_code} - Could not fetch employee for password update")

                    except Exception as pwd_error:
                        # Don't fail the entire creation if password update fails
                        print(f"   [⚠] [{i}/{len(employee_list)}] {emp_code} - Password update failed: {str(pwd_error)[:100]}")
                        # Status remains SUCCESS since employee was created

            except requests.exceptions.HTTPError as e:
                status_code = e.response.status_code if hasattr(e, 'response') and e.response is not None else 500
                error_text = e.response.text if hasattr(e, 'response') and e.response is not None else str(e)

                # Extract clean error message from API response
                error_message = self._extract_error_message(error_text) if error_text else str(e)[:200]

                # Fail fast on auth errors
                if status_code == 401:
                    print(f"\n   ❌ AUTHORIZATION FAILED - Cannot create employees")
                    print(f"   The endpoint /egov-hrms/employees/_create requires authentication.")
                    print(f"   Ask admin to add it to EGOV_OPEN_ENDPOINTS_WHITELIST.")
                    results['failed'] = len(employee_list)
                    results['errors'].append({'error': 'Authorization failed (401) - endpoint not in whitelist'})
                    return results

                # Check for duplicate/already exists
                if ('already exists' in error_text.lower() or
                    'duplicate' in error_text.lower() or
                    'user_username_unique_key' in error_text.lower() or
                    'eg_user_username_key' in error_text.lower()):
                    print(f"   [EXISTS] [{i}/{len(employee_list)}] {emp_code} (HTTP {status_code})")
                    results['exists'] += 1
                    status = "EXISTS"
                else:
                    print(f"   [FAILED] [{i}/{len(employee_list)}] {emp_code} (HTTP {status_code})")
                    print(f"   ERROR: {error_message}")
                    results['failed'] += 1
                    results['errors'].append({'id': emp_code, 'error': error_message})
                    status = "FAILED"

            except Exception as e:
                error_message = str(e)[:200]
                status_code = 0
                status = "FAILED"
                print(f"   [ERROR] [{i}/{len(employee_list)}] {emp_code} - {error_message[:100]}")
                results['failed'] += 1
                results['errors'].append({'id': emp_code, 'error': error_message})

            # Store status for this row
            row_statuses.append({
                'row_index': i,
                'status': status,
                'status_code': status_code,
                'error_message': error_message
            })

            time.sleep(0.2)

        # Summary
        print("="*60)
        print(f"[SUMMARY] Created: {results['created']}")
        print(f"[SUMMARY] Password Updated: {results['password_updated']}")
        print(f"[SUMMARY] Already Exists: {results['exists']}")
        print(f"[SUMMARY] Failed: {results['failed']}")
        print("="*60)

        # Write status columns directly into the uploaded Excel file
        if excel_file and sheet_name and row_statuses:
            self._write_status_to_excel(
                excel_file=excel_file,
                sheet_name=sheet_name,
                row_statuses=row_statuses,
                schema_code='hrms.employees'
            )

        return results

    # =========================================================================
    # WORKFLOW SERVICE METHODS
    # =========================================================================

    def search_workflow(self, tenant: str, business_service: str = "PGR") -> Dict:
        """Search for workflow business service configuration

        Args:
            tenant: Tenant ID (e.g., 'statea')
            business_service: Business service code (default: 'PGR')

        Returns:
            dict: BusinessService object if found, empty dict if not found
        """
        url = f"{self.workflow_url}/egov-wf/businessservice/_search"
        params = {
            "tenantId": tenant,
            "businessServices": business_service
        }

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "authToken": self.auth_token,
                "userInfo": self.user_info,
                "msgId": f"{int(time.time() * 1000)}|en_IN"
            }
        }

        headers = {"Content-Type": "application/json"}

        try:
            response = requests.post(url, json=payload, headers=headers, params=params)
            response.raise_for_status()

            data = response.json()
            services = data.get('BusinessServices', [])

            if services:
                return services[0]
            return {}

        except requests.exceptions.HTTPError as e:
            print(f"   Workflow search error: {e.response.status_code}")
            return {}
        except Exception as e:
            print(f"   Workflow search error: {str(e)}")
            return {}

    def create_workflow(self, tenant: str, business_service: Dict) -> Dict:
        """Create a new workflow business service

        Args:
            tenant: Tenant ID
            business_service: BusinessService object with states and actions

        Returns:
            dict: {created: bool, error: str or None, data: response data}
        """
        url = f"{self.workflow_url}/egov-wf/businessservice/_create"

        # Ensure tenantId is set on the business service
        business_service['tenantId'] = tenant

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "action": "",
                "did": 1,
                "key": "",
                "msgId": f"{int(time.time() * 1000)}|en_IN",
                "requesterId": "",
                "ts": int(time.time() * 1000),
                "ver": ".01",
                "authToken": self.auth_token,
                "userInfo": self.user_info
            },
            "BusinessServices": [business_service]
        }

        headers = {"Content-Type": "application/json"}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()

            data = response.json()
            return {
                'created': True,
                'error': None,
                'data': data.get('BusinessServices', [{}])[0]
            }

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e, 'response') and e.response else str(e)
            return {
                'created': False,
                'error': self._extract_error_message(error_text),
                'data': None
            }
        except Exception as e:
            return {
                'created': False,
                'error': str(e),
                'data': None
            }

    def update_workflow(self, tenant: str, business_service: Dict) -> Dict:
        """Update an existing workflow business service

        Args:
            tenant: Tenant ID
            business_service: BusinessService object with states and actions

        Returns:
            dict: {updated: bool, error: str or None, data: response data}
        """
        url = f"{self.workflow_url}/egov-wf/businessservice/_update"

        # Ensure tenantId is set on the business service
        business_service['tenantId'] = tenant

        payload = {
            "RequestInfo": {
                "apiId": "Rainmaker",
                "action": "",
                "did": 1,
                "key": "",
                "msgId": f"{int(time.time() * 1000)}|en_IN",
                "requesterId": "",
                "ts": int(time.time() * 1000),
                "ver": ".01",
                "authToken": self.auth_token,
                "userInfo": self.user_info
            },
            "BusinessServices": [business_service]
        }

        headers = {"Content-Type": "application/json"}

        try:
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()

            data = response.json()
            return {
                'updated': True,
                'error': None,
                'data': data.get('BusinessServices', [{}])[0]
            }

        except requests.exceptions.HTTPError as e:
            error_text = e.response.text if hasattr(e, 'response') and e.response else str(e)
            return {
                'updated': False,
                'error': self._extract_error_message(error_text),
                'data': None
            }
        except Exception as e:
            return {
                'updated': False,
                'error': str(e),
                'data': None
            }

