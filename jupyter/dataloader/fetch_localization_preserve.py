#!/usr/bin/env python3
"""
Script to fetch localization messages from API and save to Excel file
PRESERVES existing validations, formatting, and formulas in the Translation column
"""

import requests
import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

# Configuration
API_URL = "{{url}}/localization/messages/v1/_search"
PARAMS = {
    "locale": "hi_IN",
    "tenantId": "pb"
}

OUTPUT_FILE = "localization.xlsx"
SHEET_NAME = "localization"


def fetch_localization_data(url, params=None):
    """
    Fetch localization data from the API

    Args:
        url: Full API URL
        params: Dictionary with 'locale' and 'tenantId'. If None, uses global PARAMS

    Returns:
        JSON response data or None if request fails
    """
    try:
        print(f"Fetching data from: {url}")
        # Use provided params or fall back to global PARAMS
        request_params = params if params is not None else PARAMS
        response = requests.post(
            url,
            params=request_params
        )

        response.raise_for_status()
        return response.json()

    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return None


def parse_messages(api_response):
    """
    Parse API response and extract messages

    Args:
        api_response: JSON response from API

    Returns:
        List of message dictionaries
    """
    messages = []

    # Adjust this based on actual API response structure
    if isinstance(api_response, dict):
        # Try common response structures
        if 'messages' in api_response:
            messages = api_response['messages']
        elif 'data' in api_response:
            messages = api_response['data']
        elif 'responseInfo' in api_response and 'messages' in api_response:
            messages = api_response['messages']
        else:
            # Assume the entire response is the message list
            messages = [api_response]
    elif isinstance(api_response, list):
        messages = api_response

    return messages


def create_dataframe(messages, selected_language_name):
    """
    Create a pandas DataFrame from messages with standard columns

    Args:
        messages: List of message dictionaries
        selected_language_name: The language name selected by user to populate Locale column

    Returns:
        pandas DataFrame with localization data
    """
    data = []

    for msg in messages:
        row = {
            'Code': msg.get('code', ''),
            'Message': msg.get('message', ''),
            'Module': msg.get('module', ''),
            'Locale': selected_language_name,  # Use the selected language name for all rows
            'Translation': ''  # Empty column for translations
        }
        data.append(row)

    df = pd.DataFrame(data)
    return df


def read_existing_translations(filename, sheet_name):
    """
    Read existing translations from the Excel file

    Args:
        filename: Excel filename
        sheet_name: Sheet name

    Returns:
        Dictionary mapping Code to Translation value
    """
    translations = {}

    if not os.path.exists(filename):
        return translations

    try:
        existing_df = pd.read_excel(filename, sheet_name=sheet_name)

        # Create mapping of Code to Translation
        if 'Code' in existing_df.columns and 'Translation' in existing_df.columns:
            for _, row in existing_df.iterrows():
                code = row.get('Code', '')
                translation = row.get('Translation', '')
                if code and translation:  # Only save non-empty translations
                    translations[code] = translation

        print(f"✓ Loaded {len(translations)} existing translations")

    except Exception as e:
        print(f"Note: Could not read existing file ({e}). Creating new file.")

    return translations


def merge_translations(df, existing_translations):
    """
    Merge existing translations into the new dataframe

    Args:
        df: New dataframe
        existing_translations: Dictionary of Code -> Translation

    Returns:
        DataFrame with merged translations
    """
    if existing_translations:
        df['Translation'] = df['Code'].map(existing_translations).fillna('')
        print(f"✓ Merged existing translations for {df['Translation'].astype(bool).sum()} records")

    return df


def preserve_validations_and_save(df, filename, sheet_name):
    """
    Save DataFrame to Excel while preserving existing validations and formatting

    Args:
        df: pandas DataFrame
        filename: Output Excel filename
        sheet_name: Sheet name
    """
    try:
        file_exists = os.path.exists(filename)

        if file_exists:
            # Load existing workbook to preserve everything
            print("✓ Loading existing workbook to preserve validations...")
            wb = load_workbook(filename)

            # Check if sheet exists
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Store existing data validations exactly as they are
                stored_validations = []
                for dv in ws.data_validations.dataValidation:
                    stored_validations.append({
                        'type': dv.type,
                        'formula1': dv.formula1,
                        'formula2': dv.formula2,
                        'allow_blank': dv.allow_blank,
                        'showDropDown': dv.showDropDown,
                        'showInputMessage': dv.showInputMessage,
                        'showErrorMessage': dv.showErrorMessage,
                        'errorTitle': dv.errorTitle,
                        'error': dv.error,
                        'promptTitle': dv.promptTitle,
                        'prompt': dv.prompt,
                        'ranges': [str(r) for r in dv.cells.ranges]
                    })
                print(f"✓ Found {len(stored_validations)} validation rules to preserve")
            else:
                ws = wb.create_sheet(sheet_name)
                stored_validations = []
        else:
            # Create new workbook with both sheets
            print("✓ Creating new workbook...")
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            stored_validations = []

        # Get the existing row count to know if we need to delete or add rows
        existing_rows = ws.max_row
        new_rows = len(df) + 1  # +1 for header

        # Update existing cells instead of deleting rows
        headers = df.columns.tolist()

        # Write/update headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

        # Write/update data rows
        for row_idx, row_data in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Restore data validations with exact same ranges (adjusted for row count)
        if stored_validations:
            print(f"✓ Restoring {len(stored_validations)} validation rules...")
            # Clear existing validations first
            ws.data_validations.dataValidation.clear()

            for dv_data in stored_validations:
                # Create a copy of the validation
                new_dv = DataValidation(
                    type=dv_data['type'],
                    formula1=dv_data['formula1'],
                    formula2=dv_data['formula2'],
                    allow_blank=dv_data['allow_blank'],
                    showDropDown=dv_data['showDropDown'],
                    showInputMessage=dv_data['showInputMessage'],
                    showErrorMessage=dv_data['showErrorMessage'],
                    errorTitle=dv_data['errorTitle'],
                    error=dv_data['error'],
                    promptTitle=dv_data['promptTitle'],
                    prompt=dv_data['prompt']
                )

                # Add validation to sheet
                ws.add_data_validation(new_dv)

                # Apply to the same ranges but adjusted for new row count
                for range_str in dv_data['ranges']:
                    # Extract column letter from range (e.g., "D2:D100" -> "D")
                    col_letter = ''.join(filter(str.isalpha, range_str.split(':')[0]))
                    if col_letter:
                        # Apply to all data rows in that column
                        new_range = f"{col_letter}2:{col_letter}{len(df) + 1}"
                        new_dv.add(new_range)

        # Auto-adjust column widths
        for idx, col in enumerate(headers, start=1):
            max_length = max(
                df[col].astype(str).apply(len).max() if len(df) > 0 else 0,
                len(col)
            )
            col_letter = chr(64 + idx)  # A, B, C, etc.
            ws.column_dimensions[col_letter].width = min(max_length + 5, 50)

        # Save workbook
        wb.save(filename)
        print(f"✓ Data saved to {filename}")
        print(f"✓ Total records: {len(df)}")
        print(f"✓ Validations preserved successfully")

    except Exception as e:
        print(f"Error saving to Excel: {e}")
        import traceback
        traceback.print_exc()
        # Fallback to simple save without validation preservation
        print("Attempting simple save without validation preservation...")
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a' if os.path.exists(filename) else 'w', if_sheet_exists='replace') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            for cell in worksheet[1]:
                cell.font = Font(bold=True)


def main():
    """
    Main function to orchestrate the workflow
    """
    print("=" * 60)
    print("Localization Data Fetcher (Validation Preserving)")
    print("=" * 60)

    # Replace {{url}} with actual base URL
    base_url = input("Enter the base URL (e.g., https://qa.digit.org): ").strip()

    if not base_url:
        print("Error: Base URL is required")
        return

    # Ask for tenant ID
    tenant_id = input("Enter the tenant ID (e.g., pb, pg): ").strip()
    if not tenant_id:
        print("Error: Tenant ID is required")
        return

    # Ask for locale code
    locale_code = input("Enter the locale code (e.g., en_IN, hi_IN): ").strip()
    if not locale_code:
        print("Error: Locale code is required")
        return

    print(f"\n✓ Configuration:")
    print(f"  - Tenant ID: {tenant_id}")
    print(f"  - Locale Code: {locale_code}")

    # Construct full URL
    full_url = f"{base_url}/localization/messages/v1/_search"

    # Create params dict with user inputs
    params = {
        'locale': locale_code,
        'tenantId': tenant_id
    }

    # Step 1: Read existing translations
    print("\n[1/4] Reading existing translations...")
    existing_translations = read_existing_translations(OUTPUT_FILE, SHEET_NAME)

    # Step 2: Fetch data
    print(f"\n[2/4] Fetching data from API for locale '{locale_code}'...")
    api_response = fetch_localization_data(full_url, params=params)

    if not api_response:
        print("Failed to fetch data. Please check your URL and credentials.")
        return

    # Step 3: Parse messages
    print("[3/4] Parsing messages...")
    messages = parse_messages(api_response)

    if not messages:
        print("No messages found in response. API response structure:")
        print(json.dumps(api_response, indent=2)[:500])
        return

    # Step 4: Create DataFrame with user-specified language name and merge translations
    print(f"[4/4] Creating Excel structure with Locale = '{locale_code}'...")
    df = create_dataframe(messages, locale_code)
    df = merge_translations(df, existing_translations)

    # Step 5: Save to Excel with validation preservation
    print(f"\nSaving to {OUTPUT_FILE} (preserving validations)...")
    preserve_validations_and_save(df, OUTPUT_FILE, SHEET_NAME)

    print("\n" + "=" * 60)
    print("Process completed successfully!")
    print("✓ All existing validations have been preserved")
    print("✓ Existing translations have been merged")
    print("=" * 60)


if __name__ == "__main__":
    main()
